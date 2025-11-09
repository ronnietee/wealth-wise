"""
User API routes.
"""

from flask import Blueprint, request, jsonify
from ...auth import token_required, get_current_user
from ...services import UserService, EmailService
from ...utils.currency import get_currency_symbol
from ...extensions import db, limiter

user_bp = Blueprint('user', __name__, url_prefix='/user')


@user_bp.route('/profile')
@limiter.exempt  # Exempt GET requests from rate limiting
@token_required
def get_user_profile(current_user):
    """Get current user profile."""
    
    return jsonify({
        'id': current_user.id,
        'username': current_user.username,
        'email': current_user.email,
        'first_name': current_user.first_name,
        'last_name': current_user.last_name,
        'display_name': current_user.display_name,
        'preferred_name': current_user.preferred_name,
        'currency': current_user.currency,
        'theme': current_user.theme,
        'email_verified': current_user.email_verified,
        'created_at': current_user.created_at.isoformat()
    }), 200


@user_bp.route('/settings', methods=['GET'])
@limiter.exempt  # Exempt GET requests from rate limiting
@token_required
def get_user_settings(current_user):
    """Get user settings."""
    return jsonify({
        'username': current_user.username,
        'email': current_user.email,
        'first_name': current_user.first_name,
        'last_name': current_user.last_name,
        'display_name': current_user.display_name,
        'currency': current_user.currency,
        'theme': current_user.theme,
        'email_verified': current_user.email_verified
    }), 200


@user_bp.route('/settings', methods=['PUT'])
@token_required
def update_user_settings(current_user):
    """Update user settings."""
    data = request.get_json()
    
    # Update allowed fields
    allowed_fields = ['first_name', 'last_name', 'display_name', 'currency', 'theme']
    update_data = {k: v for k, v in data.items() if k in allowed_fields}
    
    if not update_data:
        return jsonify({'message': 'No valid fields to update'}), 400
    
    UserService.update_user_settings(current_user, **update_data)
    
    return jsonify({'message': 'Settings updated successfully'}), 200


@user_bp.route('/change-password', methods=['POST'])
@token_required
def change_password(current_user):
    """Change user password."""
    data = request.get_json()
    old_password = data.get('old_password')
    new_password = data.get('new_password')
    
    if not old_password or not new_password:
        return jsonify({'message': 'Old and new passwords are required'}), 400
    
    if len(new_password) < 6:
        return jsonify({'message': 'New password must be at least 6 characters long'}), 400
    
    success, error = UserService.change_password(current_user, old_password, new_password)
    if not success:
        return jsonify({'message': error}), 400
    
    return jsonify({'message': 'Password changed successfully'}), 200


@user_bp.route('/reset-data', methods=['POST'])
@token_required
def reset_user_data(current_user):
    """Reset all user data except account."""
    try:
        success, error = UserService.reset_user_data(current_user)
        if not success:
            print(f"Reset data failed: {error}")
            return jsonify({'message': error or 'Failed to reset data'}), 500
        
        return jsonify({'message': 'Data reset successfully'}), 200
    except Exception as e:
        print(f"Reset data exception: {str(e)}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({'message': f'Error resetting data: {str(e)}'}), 500


@user_bp.route('/delete-account', methods=['POST'])
@token_required
def delete_user_account(current_user):
    """Delete user account."""
    data = request.get_json()
    confirm_password = data.get('password')
    
    if not confirm_password:
        return jsonify({'message': 'Password confirmation is required'}), 400
    
    # Verify password
    from werkzeug.security import check_password_hash
    if not check_password_hash(current_user.password_hash, confirm_password):
        return jsonify({'message': 'Incorrect password'}), 400
    
    success, error = UserService.delete_user(current_user)
    if not success:
        return jsonify({'message': error}), 500
    
    return jsonify({'message': 'Account deleted successfully'}), 200


@user_bp.route('/export-data', methods=['GET'])
@limiter.exempt  # Exempt GET requests from rate limiting (large file export)
@token_required
def export_user_data(current_user):
    """Export user data to Excel."""
    try:
        # Check if openpyxl is available
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({'message': 'openpyxl library not installed. Please install it with: pip install openpyxl'}), 500
        
        import io
        from datetime import datetime
        
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create Summary Sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.append(["STEWARD - Financial Data Export"])
        summary_ws.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        summary_ws.append([f"User: {current_user.username} ({current_user.email})"])
        summary_ws.append([f"Currency: {current_user.currency}"])
        summary_ws.append([])
        
        # Get active budget period for summary
        from ...models import BudgetPeriod, Budget, BudgetAllocation, Category, Subcategory, Transaction
        
        active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
        if active_period:
            budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
            if budget:
                # Calculate total income matching dashboard calculation (income sources + balance brought forward)
                total_income_from_sources = sum(source.amount for source in budget.income_sources)
                balance_brought_forward = budget.balance_brought_forward or 0
                total_income = total_income_from_sources + balance_brought_forward
                
                # Calculate total allocated and spent amounts
                total_allocated = sum(allocation.allocated_amount for allocation in budget.allocations)
                
                # Calculate total spent from transactions within the active period
                user_subcategory_ids = db.session.query(Subcategory.id).join(Category).filter(Category.user_id == current_user.id).subquery()
                total_spent = db.session.query(db.func.sum(Transaction.amount)).filter(
                    Transaction.subcategory_id.in_(user_subcategory_ids),
                    Transaction.transaction_date >= active_period.start_date,
                    Transaction.transaction_date <= active_period.end_date
                ).scalar() or 0
                
                # Calculate remaining amounts
                remaining_to_allocate = total_income - total_allocated
                current_balance = total_income - total_spent
                
                # Calculate days remaining
                today = datetime.now().date()
                start_date = active_period.start_date
                end_date = active_period.end_date
                
                if today >= start_date and today <= end_date:
                    days_remaining = (end_date - today).days
                elif today < start_date:
                    days_remaining = (end_date - start_date).days + 1
                else:
                    days_remaining = 0
                
                # Add comprehensive dashboard-style summary
                summary_ws.append(["DASHBOARD SUMMARY"])
                summary_ws.append([f"Period: {active_period.name}"])
                summary_ws.append([f"Start Date: {active_period.start_date.strftime('%Y-%m-%d')}"])
                summary_ws.append([f"End Date: {active_period.end_date.strftime('%Y-%m-%d')}"])
                summary_ws.append([f"Days Remaining: {days_remaining}"])
                summary_ws.append([])
                
                summary_ws.append(["FINANCIAL OVERVIEW"])
                summary_ws.append([f"Total Income (Sources): {get_currency_symbol(current_user.currency)}{total_income_from_sources:,.2f}"])
                summary_ws.append([f"Balance Brought Forward: {get_currency_symbol(current_user.currency)}{balance_brought_forward:,.2f}"])
                summary_ws.append([f"TOTAL INCOME: {get_currency_symbol(current_user.currency)}{total_income:,.2f}"])
                summary_ws.append([f"Total Allocated: {get_currency_symbol(current_user.currency)}{total_allocated:,.2f}"])
                summary_ws.append([f"Total Spent: {get_currency_symbol(current_user.currency)}{total_spent:,.2f}"])
                summary_ws.append([f"Available to Allocate: {get_currency_symbol(current_user.currency)}{remaining_to_allocate:,.2f}"])
                summary_ws.append([f"Current Balance: {get_currency_symbol(current_user.currency)}{current_balance:,.2f}"])
                summary_ws.append([])
                
                # Add spending progress
                spending_percentage = (total_spent / total_income * 100) if total_income > 0 else 0
                summary_ws.append(["SPENDING PROGRESS"])
                summary_ws.append([f"Spending Percentage: {spending_percentage:.1f}%"])
                summary_ws.append([f"Budget Health: {'Good' if spending_percentage < 75 else 'Moderate' if spending_percentage < 90 else 'High Spending'}"])
                summary_ws.append([])
        
        # Get comprehensive transaction summary
        all_transactions = Transaction.query.filter_by(user_id=current_user.id).all()
        total_all_transactions = len(all_transactions)
        total_all_spent = sum(t.amount for t in all_transactions)
        
        # Get transactions for current period only
        current_period_transactions = []
        if active_period:
            current_period_transactions = Transaction.query.filter_by(user_id=current_user.id).filter(
                Transaction.transaction_date >= active_period.start_date,
                Transaction.transaction_date <= active_period.end_date
            ).all()
        
        summary_ws.append(["TRANSACTION SUMMARY"])
        summary_ws.append([f"Total Transactions (All Time): {total_all_transactions}"])
        summary_ws.append([f"Total Spent (All Time): {get_currency_symbol(current_user.currency)}{total_all_spent:,.2f}"])
        if active_period:
            summary_ws.append([f"Transactions This Period: {len(current_period_transactions)}"])
            summary_ws.append([f"Spent This Period: {get_currency_symbol(current_user.currency)}{sum(t.amount for t in current_period_transactions):,.2f}"])
        else:
            summary_ws.append(["No Active Budget Period"])
            summary_ws.append(["Please create a budget period to track spending"])
        summary_ws.append([])
        
        # Add note if no active budget period
        if not active_period:
            summary_ws.append(["NOTE"])
            summary_ws.append(["No active budget period found for this user."])
            summary_ws.append(["Dashboard calculations require an active budget period."])
            summary_ws.append(["Please create a budget period to see complete financial overview."])
            summary_ws.append([])
        
        # Create Allocations Sheet
        allocations_ws = wb.create_sheet("Allocations")
        allocations_ws.append(["Category", "Subcategory", "Allocated Amount", "Spent Amount", "Remaining", "Period"])
        
        # Style header row
        for cell in allocations_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Get all allocations with spending data
        categories = Category.query.filter_by(user_id=current_user.id).all()
        for category in categories:
            for subcategory in category.subcategories:
                # Get allocation for active period
                allocated = 0
                spent = 0
                period_name = "No Active Period"
                
                if active_period:
                    budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
                    if budget:
                        allocation = BudgetAllocation.query.filter_by(
                            budget_id=budget.id,
                            subcategory_id=subcategory.id
                        ).first()
                        if allocation:
                            allocated = allocation.allocated_amount
                        period_name = active_period.name
                        
                        # Calculate spent amount
                        spent_transactions = Transaction.query.filter_by(
                            user_id=current_user.id,
                            subcategory_id=subcategory.id
                        ).filter(
                            Transaction.transaction_date >= active_period.start_date,
                            Transaction.transaction_date <= active_period.end_date
                        ).all()
                        spent = sum(t.amount for t in spent_transactions)
                
                remaining = allocated - spent
                allocations_ws.append([
                    category.name,
                    subcategory.name,
                    allocated,
                    spent,
                    remaining,
                    period_name
                ])
        
        # Style allocation data rows
        for row in allocations_ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column in [3, 4, 5]:  # Amount columns
                    cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for column in allocations_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            allocations_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create Transactions Sheet
        transactions_ws = wb.create_sheet("Transactions")
        transactions_ws.append(["Date", "Category", "Subcategory", "Amount", "Description", "Comment"])
        
        # Style header row
        for cell in transactions_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Add transaction data
        for transaction in all_transactions:
            subcategory = Subcategory.query.get(transaction.subcategory_id)
            category = Category.query.get(subcategory.category_id) if subcategory else None
            
            transactions_ws.append([
                transaction.transaction_date.strftime('%Y-%m-%d') if transaction.transaction_date else '',
                category.name if category else 'Unknown',
                subcategory.name if subcategory else 'Unknown',
                transaction.amount,
                transaction.description or '',
                transaction.comment or ''
            ])
        
        # Style transaction data rows
        for row in transactions_ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column == 4:  # Amount column
                    cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths for transactions
        for column in transactions_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            transactions_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file
        filename = f"steward-export-{current_user.username}-{datetime.now().strftime('%Y%m%d')}.xlsx"
        from flask import send_file
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        import traceback
        print(f"Export error: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({'message': f'Export failed: {str(e)}'}), 500


@user_bp.route('/theme', methods=['PUT'])
@token_required
def update_user_theme(current_user):
    """Update user theme preference."""
    data = request.get_json()
    theme = data.get('theme')
    
    if theme not in ['light', 'dark']:
        return jsonify({'message': 'Invalid theme. Must be "light" or "dark"'}), 400
    
    current_user.theme = theme
    db.session.commit()
    
    return jsonify({'message': 'Theme updated successfully'}), 200
