from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from flask_cors import CORS
# # from flask_wtf.csrf import CSRFProtect, generate_csrf
from werkzeug.security import generate_password_hash, check_password_hash
import jwt
import os
import re
from datetime import datetime, timedelta
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import secrets
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
app.config['APP_NAME'] = 'STEWARD'
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///wealthwise.db')
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'true').lower() in ['true', 'on', '1']
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', 'noreply@steward.com')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)
CORS(app)
# csrf = CSRFProtect(app)

# Database Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    display_name = db.Column(db.String(100), nullable=True)
    currency = db.Column(db.String(10), default='USD')
    theme = db.Column(db.String(10), default='dark')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Onboarding fields (optional for existing users)
    country = db.Column(db.String(100), nullable=True, default=None)
    preferred_name = db.Column(db.String(100), nullable=True, default=None)
    referral_source = db.Column(db.String(100), nullable=True, default=None)
    referral_details = db.Column(db.Text, nullable=True, default=None)
    email_verified = db.Column(db.Boolean, default=False)
    
    # Relationships
    categories = db.relationship('Category', backref='user', lazy=True, cascade='all, delete-orphan')
    transactions = db.relationship('Transaction', backref='user', lazy=True, cascade='all, delete-orphan')
    budgets = db.relationship('Budget', backref='user', lazy=True, cascade='all, delete-orphan')
    budget_periods = db.relationship('BudgetPeriod', backref='user', lazy=True, cascade='all, delete-orphan')
    
    def set_password(self, password):
        """Set password hash for the user"""
        self.password_hash = generate_password_hash(password)

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    is_template = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    subcategories = db.relationship('Subcategory', backref='category', lazy=True, cascade='all, delete-orphan')

class Subcategory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Transaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    amount = db.Column(db.Float, nullable=False)
    description = db.Column(db.String(200))
    comment = db.Column(db.Text)
    subcategory_id = db.Column(db.Integer, db.ForeignKey('subcategory.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    transaction_date = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    subcategory = db.relationship('Subcategory', backref='transactions')

class BudgetPeriod(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)  # e.g., "January 2024", "Q1 2024", "Custom Budget"
    period_type = db.Column(db.String(20), nullable=False)  # 'monthly', 'quarterly', 'yearly', 'custom'
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    is_active = db.Column(db.Boolean, default=False)  # Only one active budget per user
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    budgets = db.relationship('Budget', backref='period', lazy=True, cascade='all, delete-orphan')

class Budget(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    period_id = db.Column(db.Integer, db.ForeignKey('budget_period.id'), nullable=False)
    total_income = db.Column(db.Float, default=0)
    balance_brought_forward = db.Column(db.Float, default=0)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    allocations = db.relationship('BudgetAllocation', backref='budget', lazy=True, cascade='all, delete-orphan')
    income_sources = db.relationship('IncomeSource', backref='budget', lazy=True, cascade='all, delete-orphan')

class IncomeSource(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    budget_id = db.Column(db.Integer, db.ForeignKey('budget.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class BudgetAllocation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    allocated_amount = db.Column(db.Float, default=0)
    subcategory_id = db.Column(db.Integer, db.ForeignKey('subcategory.id'), nullable=False)
    budget_id = db.Column(db.Integer, db.ForeignKey('budget.id'), nullable=False)

class PasswordResetToken(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    token = db.Column(db.String(100), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    used = db.Column(db.Boolean, default=False)

class EmailVerification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    token = db.Column(db.String(100), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    verified = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Account(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    account_type = db.Column(db.String(50), nullable=False)  # 'checking', 'savings', 'credit', 'investment', 'cash', 'other'
    bank_name = db.Column(db.String(100))
    account_number = db.Column(db.String(50))  # Last 4 digits or masked
    current_balance = db.Column(db.Float, default=0)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)

class RecurringIncomeSource(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    user = db.relationship('User', backref='recurring_income_sources')

class RecurringBudgetAllocation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    allocated_amount = db.Column(db.Float, default=0)
    subcategory_id = db.Column(db.Integer, db.ForeignKey('subcategory.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    user = db.relationship('User', backref='recurring_allocations')
    subcategory = db.relationship('Subcategory', backref='recurring_allocations')

# Authentication decorator
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'message': 'Token is missing!'}), 401
        
        try:
            if token.startswith('Bearer '):
                token = token[7:]
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            current_user = User.query.filter_by(id=data['user_id']).first()
        except:
            return jsonify({'message': 'Token is invalid!'}), 401
        
        return f(current_user, *args, **kwargs)
    return decorated

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/reset-password')
def reset_password_page():
    token = request.args.get('token')
    if not token:
        return render_template('reset_password.html', error='Invalid reset link')
    
    # Verify token is valid
    reset_token = PasswordResetToken.query.filter_by(
        token=token, 
        used=False
    ).first()
    
    if not reset_token or reset_token.expires_at < datetime.utcnow():
        return render_template('reset_password.html', error='Invalid or expired reset link')
    
    return render_template('reset_password.html', token=token)

def validate_session():
    """Validate current session and clear if invalid"""
    if 'user_id' in session and 'logged_in' in session:
        user = User.query.get(session['user_id'])
        if not user:
            # User no longer exists, clear session
            session.clear()
            return False
        return True
    return False

def get_current_user():
    """Get current user from JWT token in Authorization header or from session"""
    # Try to get token from Authorization header first (for API calls)
    token = request.headers.get('Authorization')
    if token:
        try:
            if token.startswith('Bearer '):
                token = token[7:]
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            user = User.query.get(data['user_id'])
            if user:
                return user
        except (jwt.ExpiredSignatureError, jwt.InvalidTokenError):
            pass
    
    # Fallback to session (for web interface)
    if validate_session():
        user = User.query.get(session['user_id'])
        if user:
            return user
    
    return None

@app.route('/api/user/profile')
def get_user_profile():
    user = get_current_user()
    if not user:
        return jsonify({'message': 'Authentication required'}), 401
    
    return jsonify({
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'display_name': user.display_name,
        'preferred_name': user.preferred_name
    })

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/budgets')
def budgets():
    return render_template('budgets.html')

@app.route('/income')
def income():
    return render_template('income.html')

@app.route('/breakdown')
def breakdown():
    return render_template('breakdown.html')

@app.route('/input')
def input_page():
    return render_template('input.html')



@app.route('/transactions')
def transactions():
    return render_template('transactions.html')

@app.route('/settings')
def settings():
    return render_template('settings.html')

# API Routes
@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json()
    
    # Validate required fields
    if not data.get('first_name') or not data.get('last_name'):
        return jsonify({'message': 'First name and last name are required'}), 400
    
    if User.query.filter_by(username=data['username']).first():
        return jsonify({'message': 'Username already exists'}), 400
    
    if User.query.filter_by(email=data['email']).first():
        return jsonify({'message': 'Email already exists'}), 400
    
    # Validate password meets stringent requirements
    password = data['password']
    if len(password) < 8:
        return jsonify({'message': 'Password must be at least 8 characters long'}), 400
    
    if not re.search(r'[a-z]', password):
        return jsonify({'message': 'Password must contain at least one lowercase letter'}), 400
    
    if not re.search(r'[A-Z]', password):
        return jsonify({'message': 'Password must contain at least one uppercase letter'}), 400
    
    if not re.search(r'[0-9]', password):
        return jsonify({'message': 'Password must contain at least one number'}), 400
    
    if not re.search(r'[^A-Za-z0-9]', password):
        return jsonify({'message': 'Password must contain at least one special character'}), 400
    
    user = User(
        username=data['username'],
        email=data['email'],
        password_hash=generate_password_hash(password),
        first_name=data['first_name'],
        last_name=data['last_name'],
        display_name=data.get('display_name', '').strip() or None,
        currency=data.get('currency', 'USD')
    )
    
    db.session.add(user)
    db.session.commit()
    
    # Create default categories
    create_default_categories(user.id)
    
    return jsonify({'message': 'User created successfully'}), 201

@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        print(f"Login attempt with data: {data}")
        username_or_email = data.get('username') or data.get('email')
        password = data.get('password')
        
        print(f"Looking for user with: {username_or_email}")
        
        if not username_or_email or not password:
            return jsonify({'message': 'Username/email and password are required'}), 400
        
        # Try to find user by username first, then by email
        user = User.query.filter_by(username=username_or_email).first()
        if not user:
            user = User.query.filter_by(email=username_or_email).first()
        
        print(f"User found: {user.username if user else 'None'}")
        
        if user and check_password_hash(user.password_hash, password):
            # Check if email is verified
            if not user.email_verified:
                return jsonify({
                    'message': 'Please verify your email address before logging in. Check your inbox for a verification link.',
                    'email_verification_required': True,
                    'email': user.email
                }), 403
            
            token = jwt.encode({
                'user_id': user.id,
                'exp': datetime.utcnow() + timedelta(hours=24)
            }, app.config['SECRET_KEY'], algorithm='HS256')
            
            return jsonify({
                'token': token,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'currency': user.currency
                }
            })
        
        return jsonify({'message': 'Invalid credentials'}), 401
        
    except Exception as e:
        print(f"Login error: {str(e)}")
        return jsonify({'message': 'Server error occurred'}), 500

# New frontend-compatible login route
@app.route('/login', methods=['POST'])
def frontend_login():
    """Login route for frontend form submission with session management"""
    data = request.get_json()
    username_or_email = data.get('username')
    password = data.get('password')
    remember = data.get('remember', False)
    
    if not username_or_email or not password:
        return jsonify({'success': False, 'message': 'Username/email and password are required'}), 400
    
    # Try to find user by username first, then by email
    user = User.query.filter_by(username=username_or_email).first()
    if not user:
        user = User.query.filter_by(email=username_or_email).first()
    
    if user and check_password_hash(user.password_hash, password):
        # Create session with unique identifier for this device
        import uuid
        session_id = str(uuid.uuid4())
        
        session['user_id'] = user.id
        session['username'] = user.username
        session['logged_in'] = True
        session['session_id'] = session_id
        
        # Set session expiration based on remember me
        if remember:
            # Remember for 30 days
            session.permanent = True
            app.permanent_session_lifetime = timedelta(days=30)
        else:
            # Session expires when browser closes
            session.permanent = False
        
        # Generate JWT token for API calls
        token = jwt.encode({
            'user_id': user.id,
            'exp': datetime.utcnow() + timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        
        return jsonify({
            'success': True,
            'message': 'Login successful',
            'redirect': '/dashboard',
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'currency': user.currency
            },
            'token': token
        })
    
    return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

@app.route('/api/csrf-token', methods=['GET'])
def get_csrf_token():
    """Get CSRF token for frontend forms"""
    return jsonify({'csrf_token': 'dummy_token'})  # Temporary for testing

@app.route('/api/session/validate', methods=['GET'])
def validate_user_session():
    """Validate current user session and return user info"""
    user = get_current_user()
    if user:
        return jsonify({
            'valid': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'currency': user.currency
            }
        })
    else:
        return jsonify({'valid': False, 'message': 'Session expired or invalid'}), 401

@app.route('/logout', methods=['POST'])
def logout():
    """Logout route to clear session"""
    session.clear()
    return jsonify({'success': True, 'message': 'Logged out successfully', 'redirect': '/'})

@app.route('/api/logout', methods=['POST'])
def api_logout():
    """API logout route"""
    session.clear()
    return jsonify({'success': True, 'message': 'Logged out successfully'})

@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    data = request.get_json()
    email = data.get('email')
    
    if not email:
        return jsonify({'message': 'Email is required'}), 400
    
    # Find user by email
    user = User.query.filter_by(email=email).first()
    
    if not user:
        # Don't reveal if email exists or not for security
        return jsonify({'message': 'If an account with that email exists, a password reset link has been sent.'}), 200
    
    # Generate reset token
    token = secrets.token_urlsafe(32)
    expires_at = datetime.utcnow() + timedelta(hours=1)  # Token expires in 1 hour
    
    # Invalidate any existing tokens for this user
    PasswordResetToken.query.filter_by(user_id=user.id, used=False).update({'used': True})
    
    # Create new reset token
    reset_token = PasswordResetToken(
        user_id=user.id,
        token=token,
        expires_at=expires_at
    )
    db.session.add(reset_token)
    db.session.commit()
    
    # Create reset URL
    reset_url = f"{request.host_url}reset-password?token={token}"
    
    # Send email
    subject = "Password Reset - STEWARD"
    body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background: linear-gradient(135deg, #8B4513 0%, #D2691E 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0;">
            <h1 style="margin: 0; font-size: 28px;">STEWARD</h1>
            <p style="margin: 10px 0 0 0; opacity: 0.9;">Christian Financial Management</p>
        </div>
        <div style="background: white; padding: 30px; border: 1px solid #e0e0e0; border-top: none; border-radius: 0 0 10px 10px;">
            <h2 style="color: #8B4513; margin-top: 0;">Password Reset Request</h2>
            <p>Hello {user.username},</p>
            <p>We received a request to reset your password for your STEWARD account. Click the button below to reset your password:</p>
            <div style="text-align: center; margin: 30px 0;">
                <a href="{reset_url}" style="background: linear-gradient(135deg, #8B4513 0%, #D2691E 100%); color: white; padding: 15px 30px; text-decoration: none; border-radius: 5px; display: inline-block; font-weight: bold;">Reset Password</a>
            </div>
            <p>If the button doesn't work, copy and paste this link into your browser:</p>
            <p style="word-break: break-all; color: #666; background: #f5f5f5; padding: 10px; border-radius: 5px;">{reset_url}</p>
            <p><strong>This link will expire in 1 hour.</strong></p>
            <p>If you didn't request this password reset, please ignore this email. Your password will remain unchanged.</p>
            <hr style="border: none; border-top: 1px solid #e0e0e0; margin: 30px 0;">
            <p style="color: #666; font-size: 14px; margin: 0;">Blessings,<br>The STEWARD Team</p>
        </div>
    </body>
    </html>
    """
    
    if send_email(user.email, subject, body):
        return jsonify({'message': 'If an account with that email exists, a password reset link has been sent.'}), 200
    else:
        return jsonify({'message': 'Error sending email. Please try again later.'}), 500

@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    data = request.get_json()
    token = data.get('token')
    new_password = data.get('password')
    
    if not token or not new_password:
        return jsonify({'message': 'Token and password are required'}), 400
    
    # Find valid token
    reset_token = PasswordResetToken.query.filter_by(
        token=token, 
        used=False
    ).first()
    
    if not reset_token:
        return jsonify({'message': 'Invalid or expired reset token'}), 400
    
    if reset_token.expires_at < datetime.utcnow():
        return jsonify({'message': 'Reset token has expired'}), 400
    
    # Validate new password meets stringent requirements
    if len(new_password) < 8:
        return jsonify({'message': 'Password must be at least 8 characters long'}), 400
    
    if not re.search(r'[a-z]', new_password):
        return jsonify({'message': 'Password must contain at least one lowercase letter'}), 400
    
    if not re.search(r'[A-Z]', new_password):
        return jsonify({'message': 'Password must contain at least one uppercase letter'}), 400
    
    if not re.search(r'[0-9]', new_password):
        return jsonify({'message': 'Password must contain at least one number'}), 400
    
    if not re.search(r'[^A-Za-z0-9]', new_password):
        return jsonify({'message': 'Password must contain at least one special character'}), 400
    
    # Update user password
    user = User.query.get(reset_token.user_id)
    if not user:
        return jsonify({'message': 'User not found'}), 400
    
    user.password_hash = generate_password_hash(new_password)
    
    # Mark token as used
    reset_token.used = True
    
    db.session.commit()
    
    return jsonify({'message': 'Password has been reset successfully'}), 200

@app.route('/api/categories', methods=['GET'])
@token_required
def get_categories(current_user):
    categories = Category.query.filter_by(user_id=current_user.id).all()
    result = []
    
    for category in categories:
        category_data = {
            'id': category.id,
            'name': category.name,
            'subcategories': []
        }
        
        for subcategory in category.subcategories:
            # Get active budget period
            active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
            allocated = 0
            spent = 0
            
            if active_period:
                budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
                if budget:
                    allocation = BudgetAllocation.query.filter_by(
                        budget_id=budget.id, 
                        subcategory_id=subcategory.id
                    ).first()
                    if allocation:
                        allocated = allocation.allocated_amount
                
                # Calculate spent amount for the current period
                spent_transactions = Transaction.query.filter_by(
                    user_id=current_user.id,
                    subcategory_id=subcategory.id
                ).filter(
                    Transaction.transaction_date >= active_period.start_date,
                    Transaction.transaction_date <= active_period.end_date
                ).all()
                
                spent = sum(t.amount for t in spent_transactions)
            
            subcategory_data = {
                'id': subcategory.id,
                'name': subcategory.name,
                'allocated': allocated,
                'spent': spent,
                'balance': allocated - spent
            }
            category_data['subcategories'].append(subcategory_data)
        
        result.append(category_data)
    
    return jsonify(result)

@app.route('/api/categories', methods=['POST'])
@token_required
def create_category(current_user):
    data = request.get_json()
    
    category = Category(
        name=data['name'],
        user_id=current_user.id
    )
    
    db.session.add(category)
    db.session.commit()
    
    return jsonify({'message': 'Category created successfully', 'id': category.id}), 201

@app.route('/api/subcategories', methods=['POST'])
@token_required
def create_subcategory(current_user):
    data = request.get_json()
    
    subcategory = Subcategory(
        name=data['name'],
        category_id=data['category_id']
    )
    
    db.session.add(subcategory)
    db.session.commit()
    
    return jsonify({'message': 'Subcategory created successfully', 'id': subcategory.id}), 201

@app.route('/api/categories/<int:category_id>', methods=['PUT'])
@token_required
def update_category(current_user, category_id):
    data = request.get_json()
    
    category = Category.query.filter_by(id=category_id, user_id=current_user.id).first()
    if not category:
        return jsonify({'message': 'Category not found'}), 404
    
    category.name = data['name']
    db.session.commit()
    
    return jsonify({'message': 'Category updated successfully'})

@app.route('/api/categories/<int:category_id>', methods=['DELETE'])
@token_required
def delete_category(current_user, category_id):
    category = Category.query.filter_by(id=category_id, user_id=current_user.id).first()
    if not category:
        return jsonify({'message': 'Category not found'}), 404
    
    db.session.delete(category)
    db.session.commit()
    
    return jsonify({'message': 'Category deleted successfully'})

@app.route('/api/subcategories/<int:subcategory_id>', methods=['PUT'])
@token_required
def update_subcategory(current_user, subcategory_id):
    data = request.get_json()
    
    subcategory = Subcategory.query.get(subcategory_id)
    if not subcategory:
        return jsonify({'message': 'Subcategory not found'}), 404
    
    # Check if user owns the parent category
    category = Category.query.filter_by(id=subcategory.category_id, user_id=current_user.id).first()
    if not category:
        return jsonify({'message': 'Subcategory not found'}), 404
    
    subcategory.name = data['name']
    db.session.commit()
    
    return jsonify({'message': 'Subcategory updated successfully'})

@app.route('/api/subcategories/<int:subcategory_id>', methods=['DELETE'])
@token_required
def delete_subcategory(current_user, subcategory_id):
    subcategory = Subcategory.query.get(subcategory_id)
    if not subcategory:
        return jsonify({'message': 'Subcategory not found'}), 404
    
    # Check if user owns the parent category
    category = Category.query.filter_by(id=subcategory.category_id, user_id=current_user.id).first()
    if not category:
        return jsonify({'message': 'Subcategory not found'}), 404
    
    db.session.delete(subcategory)
    db.session.commit()
    
    return jsonify({'message': 'Subcategory deleted successfully'})

@app.route('/api/transactions', methods=['POST'])
@token_required
def create_transaction(current_user):
    data = request.get_json()
    
    # Get transaction date from request or use current time
    if 'transaction_date' in data:
        transaction_date = datetime.fromisoformat(data['transaction_date'].replace('Z', '+00:00'))
    else:
        transaction_date = datetime.utcnow()
    
    # Get active budget period to validate date range
    active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
    
    if active_period:
        # Convert period dates to datetime for comparison
        period_start = datetime.combine(active_period.start_date, datetime.min.time())
        period_end = datetime.combine(active_period.end_date, datetime.max.time())
        
        # Validate transaction date is within active period
        if transaction_date.date() < active_period.start_date or transaction_date.date() > active_period.end_date:
            return jsonify({
                'error': f'Transaction date must be within the active budget period ({active_period.start_date} to {active_period.end_date})'
            }), 400
    
    transaction = Transaction(
        amount=data['amount'],
        description=data.get('description', ''),
        comment=data.get('comment', ''),
        subcategory_id=data['subcategory_id'],
        user_id=current_user.id,
        transaction_date=transaction_date
    )
    
    db.session.add(transaction)
    db.session.commit()
    
    return jsonify({'message': 'Transaction created successfully', 'id': transaction.id}), 201

# Budget Period Management Endpoints
@app.route('/api/budget-periods', methods=['GET'])
@token_required
def get_budget_periods(current_user):
    try:
        periods = BudgetPeriod.query.filter_by(user_id=current_user.id).order_by(BudgetPeriod.created_at.desc()).all()
        return jsonify([{
            'id': period.id,
            'name': period.name,
            'period_type': period.period_type,
            'start_date': period.start_date.isoformat() if period.start_date else None,
            'end_date': period.end_date.isoformat() if period.end_date else None,
            'is_active': period.is_active,
            'created_at': period.created_at.isoformat() if period.created_at else None
        } for period in periods])
    except Exception as e:
        print(f"Error in get_budget_periods: {str(e)}")
        return jsonify({'error': f'Error loading budget periods: {str(e)}'}), 500

@app.route('/api/budget-periods', methods=['POST'])
@token_required
def create_budget_period(current_user):
    data = request.get_json()
    
    start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
    
    # Validate date range
    if start_date >= end_date:
        return jsonify({'error': 'Start date must be before end date'}), 400
    
    # Check for overlapping periods of the same type
    # Allow overlaps between different period types (e.g., monthly and quarterly)
    overlapping_periods = BudgetPeriod.query.filter(
        BudgetPeriod.user_id == current_user.id,
        BudgetPeriod.period_type == data['period_type'],
        BudgetPeriod.start_date <= end_date,
        BudgetPeriod.end_date >= start_date
    ).all()
    
    if overlapping_periods:
        period_names = [p.name for p in overlapping_periods]
        return jsonify({
            'error': f'This {data["period_type"]} period overlaps with existing {data["period_type"]} periods: {", ".join(period_names)}'
        }), 400
    
    # Deactivate current active period of the same type
    BudgetPeriod.query.filter_by(
        user_id=current_user.id, 
        is_active=True,
        period_type=data['period_type']
    ).update({'is_active': False})
    
    # Create new period
    period = BudgetPeriod(
        name=data['name'],
        period_type=data['period_type'],
        start_date=start_date,
        end_date=end_date,
        user_id=current_user.id,
        is_active=True
    )
    db.session.add(period)
    db.session.commit()
    
    # Create budget for this period
    budget = Budget(
        period_id=period.id,
        user_id=current_user.id
    )
    db.session.add(budget)
    db.session.commit()
    
    # Auto-populate from recurring sources
    populate_budget_from_recurring(current_user, budget)
    
    # Create default categories if this is the first budget
    if not Category.query.filter_by(user_id=current_user.id).first():
        create_default_categories(current_user.id)
    
    return jsonify({
        'id': period.id,
        'name': period.name,
        'period_type': period.period_type,
        'start_date': period.start_date.isoformat(),
        'end_date': period.end_date.isoformat(),
        'is_active': period.is_active,
        'budget_id': budget.id
    }), 201

@app.route('/api/budget-periods/<int:period_id>/activate', methods=['POST'])
@token_required
def activate_budget_period(current_user, period_id):
    # Get the period to activate
    period = BudgetPeriod.query.filter_by(id=period_id, user_id=current_user.id).first()
    if not period:
        return jsonify({'message': 'Budget period not found'}), 404
    
    # Deactivate other active periods of the same type
    BudgetPeriod.query.filter_by(
        user_id=current_user.id, 
        is_active=True,
        period_type=period.period_type
    ).update({'is_active': False})
    
    # Activate selected period
    period.is_active = True
    db.session.commit()
    
    # Check if budget exists for this period, create if not
    budget = Budget.query.filter_by(period_id=period.id, user_id=current_user.id).first()
    if not budget:
        budget = Budget(
            period_id=period.id,
            user_id=current_user.id
        )
        db.session.add(budget)
        db.session.commit()
        
        # Auto-populate from recurring sources
        populate_budget_from_recurring(current_user, budget)
    
    return jsonify({'message': 'Budget period activated successfully'})

@app.route('/api/budget-periods/<int:period_id>', methods=['DELETE'])
@token_required
def delete_budget_period(current_user, period_id):
    # Get the period to delete
    period = BudgetPeriod.query.filter_by(id=period_id, user_id=current_user.id).first()
    if not period:
        return jsonify({'message': 'Budget period not found'}), 404
    
    # Delete all transactions that were created during this budget period
    # Transactions are linked to subcategories, so we need to find all transactions
    # that were created between the period's start and end dates
    transactions_to_delete = Transaction.query.filter(
        Transaction.user_id == current_user.id,
        Transaction.transaction_date >= period.start_date,
        Transaction.transaction_date <= period.end_date
    ).all()
    
    print(f"DEBUG: Deleting {len(transactions_to_delete)} transactions for period {period.name}")
    
    # Delete the transactions
    for transaction in transactions_to_delete:
        db.session.delete(transaction)
    
    # Delete the period (cascade will handle related budgets, allocations, income sources, etc.)
    db.session.delete(period)
    db.session.commit()
    
    return jsonify({'message': 'Budget period and all related data deleted successfully'})

@app.route('/api/budget', methods=['GET'])
@token_required
def get_budget(current_user):
    try:
        # Get active budget period
        active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
        
        if not active_period:
            return jsonify({
                'error': 'No active budget period found',
                'message': 'Please create a budget period to begin managing your finances.'
            }), 404
        
        # Get budget for this period
        budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
        
        if not budget:
            return jsonify({
                'error': 'No budget found for active period',
                'message': 'Please create a budget period to begin managing your finances.'
            }), 404
        
        # Calculate total income from income sources
        total_income_from_sources = sum(source.amount for source in budget.income_sources)
        
        return jsonify({
            'budget_id': budget.id,
            'period_id': active_period.id,
            'period_name': active_period.name,
            'period_type': active_period.period_type,
            'start_date': active_period.start_date.isoformat() if active_period.start_date else None,
            'end_date': active_period.end_date.isoformat() if active_period.end_date else None,
            'total_income': total_income_from_sources,
            'balance_brought_forward': budget.balance_brought_forward,
            'balance_to_allocate': total_income_from_sources + budget.balance_brought_forward - sum(a.allocated_amount for a in budget.allocations),
            'income_sources': [{'id': source.id, 'name': source.name, 'amount': source.amount} for source in budget.income_sources]
        })
    except Exception as e:
        print(f"Error in get_budget: {str(e)}")
        return jsonify({'error': f'Error loading budget: {str(e)}'}), 500

@app.route('/api/budget', methods=['PUT'])
@token_required
def update_budget(current_user):
    data = request.get_json()
    
    # Get active budget period
    active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
    if not active_period:
        return jsonify({'message': 'No active budget period found'}), 404
    
    budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
    if not budget:
        return jsonify({'message': 'No budget found for active period'}), 404
    
    budget.total_income = data.get('total_income', budget.total_income)
    budget.balance_brought_forward = data.get('balance_brought_forward', budget.balance_brought_forward)
    
    db.session.commit()
    
    return jsonify({'message': 'Budget updated successfully'})

@app.route('/api/allocations', methods=['POST'])
@token_required
def update_allocations(current_user):
    data = request.get_json()
    
    # Get active budget period
    active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
    if not active_period:
        return jsonify({'message': 'No active budget period found'}), 404
    
    budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
    if not budget:
        return jsonify({'message': 'No budget found for active period'}), 404
    
    # Calculate total allocation amount
    total_allocation = sum(allocation['amount'] for allocation in data['allocations'])
    available_amount = budget.total_income + budget.balance_brought_forward
    
    # Validate that total allocation doesn't exceed available amount
    if total_allocation > available_amount:
        return jsonify({'message': f'Total allocation (${total_allocation:.2f}) cannot exceed available amount (${available_amount:.2f})'}), 400
    
    # Clear existing allocations
    BudgetAllocation.query.filter_by(budget_id=budget.id).delete()
    
    # Add new allocations
    for allocation in data['allocations']:
        if allocation['amount'] > 0:  # Only add allocations with positive amounts
            new_allocation = BudgetAllocation(
                allocated_amount=allocation['amount'],
                subcategory_id=allocation['subcategory_id'],
                budget_id=budget.id
            )
            db.session.add(new_allocation)
    
    db.session.commit()
    
    return jsonify({'message': 'Allocations updated successfully'})

@app.route('/api/income-sources', methods=['POST'])
@token_required
def create_income_source(current_user):
    try:
        data = request.get_json()
        
        # Get active budget period
        active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
        
        if not active_period:
            # Create default monthly period for current month
            current_date = datetime.now()
            start_date = current_date.replace(day=1).date()
            end_date = (current_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            active_period = BudgetPeriod(
                name=current_date.strftime('%B %Y'),
                period_type='monthly',
                start_date=start_date,
                end_date=end_date,
                user_id=current_user.id,
                is_active=True
            )
            db.session.add(active_period)
            db.session.commit()
        
        budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
        
        if not budget:
            budget = Budget(
                period_id=active_period.id,
                user_id=current_user.id
            )
            db.session.add(budget)
            db.session.commit()
            
            # Auto-populate from recurring sources
            populate_budget_from_recurring(current_user, budget)
        
        income_source = IncomeSource(
            name=data['name'],
            amount=data['amount'],
            budget_id=budget.id
        )
        
        db.session.add(income_source)
        db.session.commit()
        
        # Update total income
        budget.total_income = sum(source.amount for source in budget.income_sources)
        db.session.commit()
        
        return jsonify({'message': 'Income source created successfully', 'id': income_source.id}), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error creating income source: {str(e)}'}), 500

@app.route('/api/income-sources/<int:source_id>', methods=['PUT'])
@token_required
def update_income_source(current_user, source_id):
    data = request.get_json()
    
    # Get active budget period
    active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
    if not active_period:
        return jsonify({'message': 'No active budget period found'}), 404
    
    budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
    income_source = IncomeSource.query.filter_by(id=source_id, budget_id=budget.id).first()
    
    if not income_source:
        return jsonify({'message': 'Income source not found'}), 404
    
    income_source.name = data['name']
    income_source.amount = data['amount']
    
    db.session.commit()
    
    # Update total income
    budget.total_income = sum(source.amount for source in budget.income_sources)
    db.session.commit()
    
    return jsonify({'message': 'Income source updated successfully'})

@app.route('/api/income-sources/<int:source_id>', methods=['DELETE'])
@token_required
def delete_income_source(current_user, source_id):
    # Get active budget period
    active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
    if not active_period:
        return jsonify({'message': 'No active budget period found'}), 404
    
    budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
    income_source = IncomeSource.query.filter_by(id=source_id, budget_id=budget.id).first()
    
    if not income_source:
        return jsonify({'message': 'Income source not found'}), 404
    
    db.session.delete(income_source)
    db.session.commit()
    
    # Update total income
    budget.total_income = sum(source.amount for source in budget.income_sources)
    db.session.commit()
    
    return jsonify({'message': 'Income source deleted successfully'})

@app.route('/api/user/settings', methods=['GET'])
@token_required
def get_user_settings(current_user):
    return jsonify({
        'username': current_user.username,
        'email': current_user.email,
        'first_name': current_user.first_name,
        'last_name': current_user.last_name,
        'display_name': current_user.display_name,
        'preferred_name': current_user.preferred_name,
        'currency': current_user.currency
    })

@app.route('/api/user/settings', methods=['PUT'])
@token_required
def update_user_settings(current_user):
    data = request.get_json()
    
    if 'currency' in data:
        current_user.currency = data['currency']
    
    if 'email' in data:
        # Check if email is already taken by another user
        existing_user = User.query.filter_by(email=data['email']).first()
        if existing_user and existing_user.id != current_user.id:
            return jsonify({'message': 'Email already exists'}), 400
        current_user.email = data['email']
    
    if 'display_name' in data:
        current_user.display_name = data['display_name'].strip() or None
    
    db.session.commit()
    
    return jsonify({'message': 'Settings updated successfully'})

@app.route('/api/user/change-password', methods=['POST'])
@token_required
def change_password(current_user):
    data = request.get_json()
    
    if not check_password_hash(current_user.password_hash, data['current_password']):
        return jsonify({'message': 'Current password is incorrect'}), 400
    
    # Validate new password meets stringent requirements
    new_password = data['new_password']
    if len(new_password) < 8:
        return jsonify({'message': 'Password must be at least 8 characters long'}), 400
    
    if not re.search(r'[a-z]', new_password):
        return jsonify({'message': 'Password must contain at least one lowercase letter'}), 400
    
    if not re.search(r'[A-Z]', new_password):
        return jsonify({'message': 'Password must contain at least one uppercase letter'}), 400
    
    if not re.search(r'[0-9]', new_password):
        return jsonify({'message': 'Password must contain at least one number'}), 400
    
    if not re.search(r'[^A-Za-z0-9]', new_password):
        return jsonify({'message': 'Password must contain at least one special character'}), 400
    
    current_user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    
    return jsonify({'message': 'Password changed successfully'})

@app.route('/api/user/reset-data', methods=['POST'])
@token_required
def reset_user_data(current_user):
    try:
        # Delete in correct order to respect foreign key constraints
        
        # First delete transactions (they reference subcategories)
        # Get subcategory IDs for this user first
        user_subcategory_ids = db.session.query(Subcategory.id).join(Category).filter(Category.user_id == current_user.id).subquery()
        Transaction.query.filter(Transaction.subcategory_id.in_(user_subcategory_ids)).delete(synchronize_session=False)
        
        # Delete budget allocations (they reference subcategories and budgets)
        # Get budget IDs for this user first
        user_budget_ids = db.session.query(Budget.id).filter(Budget.user_id == current_user.id).subquery()
        BudgetAllocation.query.filter(BudgetAllocation.budget_id.in_(user_budget_ids)).delete(synchronize_session=False)
        
        # Delete income sources (they reference budgets)
        IncomeSource.query.filter(IncomeSource.budget_id.in_(user_budget_ids)).delete(synchronize_session=False)
        
        # Delete budgets (they reference budget periods)
        Budget.query.filter_by(user_id=current_user.id).delete()
        
        # Delete budget periods
        BudgetPeriod.query.filter_by(user_id=current_user.id).delete()
        
        # Delete subcategories (they reference categories)
        # Get category IDs for this user first
        user_category_ids = db.session.query(Category.id).filter(Category.user_id == current_user.id).subquery()
        Subcategory.query.filter(Subcategory.category_id.in_(user_category_ids)).delete(synchronize_session=False)
        
        # Finally delete categories
        Category.query.filter_by(user_id=current_user.id).delete()
        
        db.session.commit()
        
        return jsonify({'message': 'All data has been reset successfully'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in reset_user_data: {str(e)}")
        return jsonify({'message': f'Error resetting data: {str(e)}'}), 500

@app.route('/api/user/delete-account', methods=['POST'])
@token_required
def delete_user_account(current_user):
    try:
        data = request.get_json()
        password = data.get('password')
        
        if not password:
            return jsonify({'message': 'Password is required'}), 400
        
        # Verify current password
        if not check_password_hash(current_user.password_hash, password):
            return jsonify({'message': 'Incorrect password'}), 400
        
        # Delete in correct order to respect foreign key constraints
        
        # First delete transactions (they reference subcategories)
        # Get subcategory IDs for this user first
        user_subcategory_ids = db.session.query(Subcategory.id).join(Category).filter(Category.user_id == current_user.id).subquery()
        Transaction.query.filter(Transaction.subcategory_id.in_(user_subcategory_ids)).delete(synchronize_session=False)
        
        # Delete budget allocations (they reference subcategories and budgets)
        # Get budget IDs for this user first
        user_budget_ids = db.session.query(Budget.id).filter(Budget.user_id == current_user.id).subquery()
        BudgetAllocation.query.filter(BudgetAllocation.budget_id.in_(user_budget_ids)).delete(synchronize_session=False)
        
        # Delete income sources (they reference budgets)
        IncomeSource.query.filter(IncomeSource.budget_id.in_(user_budget_ids)).delete(synchronize_session=False)
        
        # Delete budgets (they reference budget periods)
        Budget.query.filter_by(user_id=current_user.id).delete()
        
        # Delete budget periods
        BudgetPeriod.query.filter_by(user_id=current_user.id).delete()
        
        # Delete subcategories (they reference categories)
        Subcategory.query.filter(Subcategory.category_id.in_(db.session.query(Category.id).filter(Category.user_id == current_user.id).subquery())).delete(synchronize_session=False)
        
        # Delete categories
        Category.query.filter_by(user_id=current_user.id).delete()
        
        # Delete password reset tokens
        PasswordResetToken.query.filter_by(user_id=current_user.id).delete()
        
        # Finally delete the user account
        User.query.filter_by(id=current_user.id).delete()
        
        db.session.commit()
        
        return jsonify({'message': 'Account deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in delete_user_account: {str(e)}")
        return jsonify({'message': f'Error deleting account: {str(e)}'}), 500

# Accounts API Routes
@app.route('/api/accounts', methods=['GET'])
@token_required
def get_accounts(current_user):
    accounts = Account.query.filter_by(user_id=current_user.id, is_active=True).all()
    return jsonify([{
        'id': account.id,
        'name': account.name,
        'account_type': account.account_type,
        'bank_name': account.bank_name,
        'account_number': account.account_number,
        'current_balance': account.current_balance,
        'created_at': account.created_at.isoformat(),
        'updated_at': account.updated_at.isoformat()
    } for account in accounts])

@app.route('/api/accounts', methods=['POST'])
@token_required
def create_account(current_user):
    data = request.get_json()
    
    # Validate required fields
    if not data.get('name') or not data.get('account_type'):
        return jsonify({'message': 'Account name and type are required'}), 400
    
    # Validate account type
    valid_types = ['checking', 'savings', 'credit', 'investment', 'cash', 'other']
    if data['account_type'] not in valid_types:
        return jsonify({'message': 'Invalid account type'}), 400
    
    try:
        account = Account(
            name=data['name'],
            account_type=data['account_type'],
            bank_name=data.get('bank_name'),
            account_number=data.get('account_number'),
            current_balance=data.get('current_balance', 0),
            user_id=current_user.id
        )
        
        db.session.add(account)
        db.session.commit()
        
        return jsonify({
            'id': account.id,
            'name': account.name,
            'account_type': account.account_type,
            'bank_name': account.bank_name,
            'account_number': account.account_number,
            'current_balance': account.current_balance,
            'created_at': account.created_at.isoformat(),
            'updated_at': account.updated_at.isoformat()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error creating account: {str(e)}'}), 500

@app.route('/api/accounts/<int:account_id>', methods=['PUT'])
@token_required
def update_account(current_user, account_id):
    account = Account.query.filter_by(id=account_id, user_id=current_user.id, is_active=True).first()
    if not account:
        return jsonify({'message': 'Account not found'}), 404
    
    data = request.get_json()
    
    # Validate account type if provided
    if 'account_type' in data:
        valid_types = ['checking', 'savings', 'credit', 'investment', 'cash', 'other']
        if data['account_type'] not in valid_types:
            return jsonify({'message': 'Invalid account type'}), 400
    
    try:
        if 'name' in data:
            account.name = data['name']
        if 'account_type' in data:
            account.account_type = data['account_type']
        if 'bank_name' in data:
            account.bank_name = data['bank_name']
        if 'account_number' in data:
            account.account_number = data['account_number']
        if 'current_balance' in data:
            account.current_balance = data['current_balance']
        
        account.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'id': account.id,
            'name': account.name,
            'account_type': account.account_type,
            'bank_name': account.bank_name,
            'account_number': account.account_number,
            'current_balance': account.current_balance,
            'created_at': account.created_at.isoformat(),
            'updated_at': account.updated_at.isoformat()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error updating account: {str(e)}'}), 500

@app.route('/api/accounts/<int:account_id>', methods=['DELETE'])
@token_required
def delete_account(current_user, account_id):
    account = Account.query.filter_by(id=account_id, user_id=current_user.id, is_active=True).first()
    if not account:
        return jsonify({'message': 'Account not found'}), 404
    
    try:
        account.is_active = False
        db.session.commit()
        return jsonify({'message': 'Account deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error deleting account: {str(e)}'}), 500

@app.route('/api/accounts/balance-summary', methods=['GET'])
@token_required
def get_balance_summary(current_user):
    try:
        # Get total of all accounts
        accounts = Account.query.filter_by(user_id=current_user.id, is_active=True).all()
        total_accounts_balance = sum(account.current_balance for account in accounts)
        
        # Get application calculated balance from active budget period
        app_balance = 0
        
        # Get active budget period
        active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
        
        if active_period:
            budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
            
            if budget:
                # Calculate total income from income sources
                total_income = sum(source.amount for source in budget.income_sources)
                total_income += budget.balance_brought_forward or 0
                
                # Calculate total spent from transactions within the active period
                user_subcategory_ids = db.session.query(Subcategory.id).join(Category).filter(Category.user_id == current_user.id).subquery()
                total_spent = db.session.query(db.func.sum(Transaction.amount)).filter(
                    Transaction.subcategory_id.in_(user_subcategory_ids),
                    Transaction.transaction_date >= active_period.start_date,
                    Transaction.transaction_date <= active_period.end_date
                ).scalar() or 0
                
                app_balance = total_income - total_spent
        
        # Calculate alignment
        balance_difference = total_accounts_balance - app_balance
        alignment_percentage = (app_balance / total_accounts_balance * 100) if total_accounts_balance != 0 else 0
        
        return jsonify({
            'total_accounts_balance': total_accounts_balance,
            'app_balance': app_balance,
            'balance_difference': balance_difference,
            'alignment_percentage': round(alignment_percentage, 2),
            'is_aligned': abs(balance_difference) < 0.01,  # Consider aligned if difference is less than 1 cent
            'accounts_count': len(accounts)
        })
        
    except Exception as e:
        print(f"Error in get_balance_summary: {str(e)}")
        return jsonify({'message': f'Error calculating balance summary: {str(e)}'}), 500

@app.route('/accounts')
def accounts_page():
    return render_template('accounts.html')

@app.route('/favicon.ico')
def favicon():
    return send_file('static/images/logo.png', mimetype='image/png')

@app.route('/contact')
def contact():
    return render_template('contact.html')


@app.route('/api/contact', methods=['POST'])
def submit_contact():
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['name', 'email', 'subject', 'message']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'message': f'{field.replace("_", " ").title()} is required'}), 400
        
        # Validate email format
        email = data['email']
        if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
            return jsonify({'message': 'Invalid email format'}), 400
        
        # Validate message length
        message = data['message']
        if len(message) < 10:
            return jsonify({'message': 'Message must be at least 10 characters long'}), 400
        if len(message) > 2000:
            return jsonify({'message': 'Message must be less than 2000 characters'}), 400
        
        # Prepare email content
        subject_map = {
            'bug-report': 'Bug Report',
            'feature-request': 'Feature Request',
            'general-inquiry': 'General Inquiry',
            'account-issue': 'Account Issue',
            'billing-question': 'Billing Question',
            'other': 'Other'
        }
        
        priority_map = {
            'low': 'Low',
            'medium': 'Medium',
            'high': 'High'
        }
        
        email_subject = f"STEWARD Contact Form: {subject_map.get(data['subject'], 'General Inquiry')} - {priority_map.get(data.get('priority', 'medium'), 'Medium')} Priority"
        
        # Create HTML email content
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: #8B4513; color: white; padding: 20px; text-align: center; border-radius: 8px 8px 0 0; }}
                .content {{ background: #f8f9fa; padding: 20px; border-radius: 0 0 8px 8px; }}
                .field {{ margin-bottom: 15px; }}
                .label {{ font-weight: bold; color: #8B4513; }}
                .value {{ margin-top: 5px; }}
                .message {{ background: white; padding: 15px; border-radius: 5px; border-left: 4px solid #8B4513; }}
                .footer {{ margin-top: 20px; padding-top: 20px; border-top: 1px solid #ddd; font-size: 0.9em; color: #666; }}
                .priority-high {{ color: #dc3545; font-weight: bold; }}
                .priority-medium {{ color: #ffc107; font-weight: bold; }}
                .priority-low {{ color: #28a745; font-weight: bold; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2>STEWARD Contact Form Submission</h2>
                </div>
                <div class="content">
                    <div class="field">
                        <div class="label">Name:</div>
                        <div class="value">{data['name']}</div>
                    </div>
                    <div class="field">
                        <div class="label">Email:</div>
                        <div class="value">{data['email']}</div>
                    </div>
                    <div class="field">
                        <div class="label">Subject:</div>
                        <div class="value">{subject_map.get(data['subject'], 'General Inquiry')}</div>
                    </div>
                    <div class="field">
                        <div class="label">Priority:</div>
                        <div class="value priority-{data.get('priority', 'medium')}">{priority_map.get(data.get('priority', 'medium'), 'Medium')}</div>
                    </div>
                    <div class="field">
                        <div class="label">Message:</div>
                        <div class="value message">{message.replace(chr(10), '<br>')}</div>
                    </div>
                    <div class="field">
                        <div class="label">Newsletter Subscription:</div>
                        <div class="value">{'Yes' if data.get('newsletter') else 'No'}</div>
                    </div>
                    <div class="footer">
                        <p><strong>Submitted:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                        <p><strong>IP Address:</strong> {request.remote_addr}</p>
                        <p>This message was sent from the STEWARD contact form.</p>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Create plain text version
        text_content = f"""
STEWARD Contact Form Submission
==============================

Name: {data['name']}
Email: {data['email']}
Subject: {subject_map.get(data['subject'], 'General Inquiry')}
Priority: {priority_map.get(data.get('priority', 'medium'), 'Medium')}

Message:
{message}

Newsletter Subscription: {'Yes' if data.get('newsletter') else 'No'}

Submitted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
IP Address: {request.remote_addr}

This message was sent from the STEWARD contact form.
        """
        
        # Send email
        admin_email = os.getenv('ADMIN_EMAIL', 'admin@steward.com')
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = email_subject
        msg['From'] = os.getenv('SMTP_USERNAME', 'noreply@steward.com')
        msg['To'] = admin_email
        msg['Reply-To'] = email
        
        # Add both plain text and HTML versions
        text_part = MIMEText(text_content, 'plain')
        html_part = MIMEText(html_content, 'html')
        
        msg.attach(text_part)
        msg.attach(html_part)
        
        # Send email
        try:
            with smtplib.SMTP(os.getenv('SMTP_SERVER', 'smtp.gmail.com'), int(os.getenv('SMTP_PORT', '587'))) as server:
                server.starttls()
                server.login(os.getenv('SMTP_USERNAME', ''), os.getenv('SMTP_PASSWORD', ''))
                server.send_message(msg)
            
            return jsonify({'message': 'Message sent successfully!'})
            
        except Exception as email_error:
            print(f"Error sending contact email: {str(email_error)}")
            return jsonify({'message': 'Message received but failed to send email notification. We will still review your message.'}), 500
        
    except Exception as e:
        print(f"Error in submit_contact: {str(e)}")
        return jsonify({'message': 'An error occurred while processing your message. Please try again.'}), 500

@app.route('/api/user/export-data', methods=['GET'])
@token_required
def export_user_data(current_user):
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create Summary Sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.append(["STEWARD - Financial Data Export"])
        summary_ws.append([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        summary_ws.append([f"User: {current_user.username} ({current_user.email})"])
        summary_ws.append([f"Currency: {current_user.currency}"])
        summary_ws.append([])
        
        # Get active budget period for summary
        active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
        if active_period:
            budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
            if budget:
                # Calculate total income matching dashboard calculation (income sources + balance brought forward)
                total_income_from_sources = sum(source.amount for source in budget.income_sources)
                balance_brought_forward = budget.balance_brought_forward or 0
                total_income = total_income_from_sources + balance_brought_forward
                
                # Calculate total allocated and spent amounts
                total_allocated = sum(allocation.allocated_amount for allocation in budget.allocations)
                
                # Calculate total spent from transactions within the active period
                user_subcategory_ids = db.session.query(Subcategory.id).join(Category).filter(Category.user_id == current_user.id).subquery()
                total_spent = db.session.query(db.func.sum(Transaction.amount)).filter(
                    Transaction.subcategory_id.in_(user_subcategory_ids),
                    Transaction.transaction_date >= active_period.start_date,
                    Transaction.transaction_date <= active_period.end_date
                ).scalar() or 0
                
                # Calculate remaining amounts
                remaining_to_allocate = total_income - total_allocated
                current_balance = total_income - total_spent
                
                # Calculate days remaining
                today = datetime.now().date()
                start_date = active_period.start_date
                end_date = active_period.end_date
                
                if today >= start_date and today <= end_date:
                    days_remaining = (end_date - today).days
                elif today < start_date:
                    days_remaining = (end_date - start_date).days + 1
                else:
                    days_remaining = 0
                
                # Add comprehensive dashboard-style summary
                summary_ws.append(["DASHBOARD SUMMARY"])
                summary_ws.append([f"Period: {active_period.name}"])
                summary_ws.append([f"Start Date: {active_period.start_date.strftime('%Y-%m-%d')}"])
                summary_ws.append([f"End Date: {active_period.end_date.strftime('%Y-%m-%d')}"])
                summary_ws.append([f"Days Remaining: {days_remaining}"])
                summary_ws.append([])
                
                summary_ws.append(["FINANCIAL OVERVIEW"])
                summary_ws.append([f"Total Income (Sources): {getCurrencySymbol(current_user.currency)}{total_income_from_sources:,.2f}"])
                summary_ws.append([f"Balance Brought Forward: {getCurrencySymbol(current_user.currency)}{balance_brought_forward:,.2f}"])
                summary_ws.append([f"TOTAL INCOME: {getCurrencySymbol(current_user.currency)}{total_income:,.2f}"])
                summary_ws.append([f"Total Allocated: {getCurrencySymbol(current_user.currency)}{total_allocated:,.2f}"])
                summary_ws.append([f"Total Spent: {getCurrencySymbol(current_user.currency)}{total_spent:,.2f}"])
                summary_ws.append([f"Available to Allocate: {getCurrencySymbol(current_user.currency)}{remaining_to_allocate:,.2f}"])
                summary_ws.append([f"Current Balance: {getCurrencySymbol(current_user.currency)}{current_balance:,.2f}"])
                summary_ws.append([])
                
                # Add spending progress
                spending_percentage = (total_spent / total_income * 100) if total_income > 0 else 0
                summary_ws.append(["SPENDING PROGRESS"])
                summary_ws.append([f"Spending Percentage: {spending_percentage:.1f}%"])
                summary_ws.append([f"Budget Health: {'Good' if spending_percentage < 75 else 'Moderate' if spending_percentage < 90 else 'High Spending'}"])
                summary_ws.append([])
        
        # Get comprehensive transaction summary
        all_transactions = Transaction.query.filter_by(user_id=current_user.id).all()
        total_all_transactions = len(all_transactions)
        total_all_spent = sum(t.amount for t in all_transactions)
        
        # Get transactions for current period only
        current_period_transactions = []
        if active_period:
            current_period_transactions = Transaction.query.filter_by(user_id=current_user.id).filter(
                Transaction.transaction_date >= active_period.start_date,
                Transaction.transaction_date <= active_period.end_date
            ).all()
        
        summary_ws.append(["TRANSACTION SUMMARY"])
        summary_ws.append([f"Total Transactions (All Time): {total_all_transactions}"])
        summary_ws.append([f"Total Spent (All Time): {getCurrencySymbol(current_user.currency)}{total_all_spent:,.2f}"])
        if active_period:
            summary_ws.append([f"Transactions This Period: {len(current_period_transactions)}"])
            summary_ws.append([f"Spent This Period: {getCurrencySymbol(current_user.currency)}{sum(t.amount for t in current_period_transactions):,.2f}"])
        else:
            summary_ws.append(["No Active Budget Period"])
            summary_ws.append(["Please create a budget period to track spending"])
        summary_ws.append([])
        
        # Add note if no active budget period
        if not active_period:
            summary_ws.append(["NOTE"])
            summary_ws.append(["No active budget period found for this user."])
            summary_ws.append(["Dashboard calculations require an active budget period."])
            summary_ws.append(["Please create a budget period to see complete financial overview."])
            summary_ws.append([])
        
        # Create Allocations Sheet
        allocations_ws = wb.create_sheet("Allocations")
        allocations_ws.append(["Category", "Subcategory", "Allocated Amount", "Spent Amount", "Remaining", "Period"])
        
        # Style header row
        for cell in allocations_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Get all allocations with spending data
        categories = Category.query.filter_by(user_id=current_user.id).all()
        for category in categories:
            for subcategory in category.subcategories:
                # Get allocation for active period
                allocated = 0
                spent = 0
                period_name = "No Active Period"
                
                if active_period:
                    budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
                    if budget:
                        allocation = BudgetAllocation.query.filter_by(
                            budget_id=budget.id,
                            subcategory_id=subcategory.id
                        ).first()
                        if allocation:
                            allocated = allocation.allocated_amount
                        period_name = active_period.name
                        
                        # Calculate spent amount
                        spent_transactions = Transaction.query.filter_by(
                            user_id=current_user.id,
                            subcategory_id=subcategory.id
                        ).filter(
                            Transaction.transaction_date >= active_period.start_date,
                            Transaction.transaction_date <= active_period.end_date
                        ).all()
                        spent = sum(t.amount for t in spent_transactions)
                
                remaining = allocated - spent
                allocations_ws.append([
                    category.name,
                    subcategory.name,
                    allocated,
                    spent,
                    remaining,
                    period_name
                ])
        
        # Style allocation data rows
        for row in allocations_ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column in [3, 4, 5]:  # Amount columns
                    cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for column in allocations_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            allocations_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create Transactions Sheet
        transactions_ws = wb.create_sheet("Transactions")
        transactions_ws.append(["Date", "Category", "Subcategory", "Amount", "Description", "Comment"])
        
        # Style header row
        for cell in transactions_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Add transaction data
        for transaction in all_transactions:
            subcategory = Subcategory.query.get(transaction.subcategory_id)
            category = Category.query.get(subcategory.category_id) if subcategory else None
            
            transactions_ws.append([
                transaction.transaction_date.strftime('%Y-%m-%d') if transaction.transaction_date else '',
                category.name if category else 'Unknown',
                subcategory.name if subcategory else 'Unknown',
                transaction.amount,
                transaction.description or '',
                transaction.comment or ''
            ])
        
        # Style transaction data rows
        for row in transactions_ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column == 4:  # Amount column
                    cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths for transactions
        for column in transactions_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            transactions_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return file
        filename = f"steward-export-{current_user.username}-{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error in export_user_data: {str(e)}")
        return jsonify({'message': f'Error exporting data: {str(e)}'}), 500

def getCurrencySymbol(currency):
    symbols = {
        'USD': '$', 'EUR': '€', 'GBP': '£', 'ZAR': 'R', 'CAD': 'C$', 'AUD': 'A$',
        'BWP': 'P', 'ZMW': 'K', 'NGN': '₦', 'KES': 'KSh', 'GHS': '₵', 'UGX': 'USh',
        'TZS': 'TSh', 'ETB': 'Br', 'RWF': 'RF', 'MWK': 'MK', 'BRL': 'R$', 'MXN': '$',
        'PHP': '₱', 'INR': '₹', 'JPY': '¥'
    }
    return symbols.get(currency, '$')

@app.route('/api/transactions', methods=['GET'])
@token_required
def get_transactions(current_user):
    try:
        # Get period type from query parameter (default to 'monthly')
        period_type = request.args.get('period_type', 'monthly')
        
        # Get active budget period of the specified type
        active_period = BudgetPeriod.query.filter_by(
            user_id=current_user.id, 
            is_active=True,
            period_type=period_type
        ).first()
        
        if not active_period:
            return jsonify([])
        
        # Filter transactions by active budget period date range
        transactions = Transaction.query.filter_by(user_id=current_user.id).filter(
            Transaction.transaction_date >= active_period.start_date,
            Transaction.transaction_date <= active_period.end_date
        ).order_by(Transaction.transaction_date.desc()).all()
        
        result = []
        for transaction in transactions:
            subcategory = Subcategory.query.get(transaction.subcategory_id)
            category = Category.query.get(subcategory.category_id) if subcategory else None
            
            result.append({
                'id': transaction.id,
                'amount': transaction.amount,
                'description': transaction.description or '',
                'comment': transaction.comment or '',
                'subcategory_id': transaction.subcategory_id,
                'transaction_date': transaction.transaction_date.isoformat(),
                'category_name': category.name if category else 'Unknown',
                'subcategory_name': subcategory.name if subcategory else 'Unknown'
            })
        
        return jsonify(result)
    except Exception as e:
        print(f"Error in get_transactions: {str(e)}")
        return jsonify({'error': f'Error loading transactions: {str(e)}'}), 500

@app.route('/api/transactions/<int:transaction_id>', methods=['PUT'])
@token_required
def update_transaction(current_user, transaction_id):
    transaction = Transaction.query.filter_by(id=transaction_id, user_id=current_user.id).first()
    
    if not transaction:
        return jsonify({'message': 'Transaction not found'}), 404
    
    data = request.get_json()
    
    # Check if transaction_date is being updated
    if 'transaction_date' in data:
        new_transaction_date = datetime.fromisoformat(data['transaction_date'].replace('Z', '+00:00'))
        
        # Get active budget period to validate date range
        active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
        
        if active_period:
            # Validate new transaction date is within active period
            if new_transaction_date.date() < active_period.start_date or new_transaction_date.date() > active_period.end_date:
                return jsonify({
                    'error': f'Transaction date must be within the active budget period ({active_period.start_date} to {active_period.end_date})'
                }), 400
        
        transaction.transaction_date = new_transaction_date
    
    transaction.amount = data.get('amount', transaction.amount)
    transaction.description = data.get('description', transaction.description)
    transaction.comment = data.get('comment', transaction.comment)
    transaction.subcategory_id = data.get('subcategory_id', transaction.subcategory_id)
    
    db.session.commit()
    
    return jsonify({'message': 'Transaction updated successfully'})

@app.route('/api/transactions/<int:transaction_id>', methods=['DELETE'])
@token_required
def delete_transaction(current_user, transaction_id):
    transaction = Transaction.query.filter_by(id=transaction_id, user_id=current_user.id).first()
    
    if not transaction:
        return jsonify({'message': 'Transaction not found'}), 404
    
    db.session.delete(transaction)
    db.session.commit()
    
    return jsonify({'message': 'Transaction deleted successfully'})

@app.route('/api/budget/balance-check', methods=['GET'])
@token_required
def check_budget_balance(current_user):
    """Check if total income is sufficient for total allocated budget"""
    try:
        # Get active budget period
        active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
        
        if not active_period:
            return jsonify({
                'is_balanced': True,
                'message': 'No active budget period found'
            })
        
        budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
        
        if not budget:
            return jsonify({
                'is_balanced': True,
                'message': 'No budget found for active period'
            })
        
        # Calculate total income from income sources
        total_income = sum(source.amount for source in budget.income_sources)
        total_income += budget.balance_brought_forward or 0
        
        # Calculate total allocated amount
        total_allocated = sum(allocation.allocated_amount for allocation in budget.allocations)
        
        # Check if budget is balanced
        is_balanced = total_income >= total_allocated
        deficit = total_allocated - total_income if not is_balanced else 0
        
        return jsonify({
            'is_balanced': is_balanced,
            'total_income': total_income,
            'total_allocated': total_allocated,
            'deficit': deficit,
            'message': f'Budget is {"balanced" if is_balanced else "unbalanced"}. {"Deficit: $" + f"{deficit:.2f}" if not is_balanced else ""}'
        })
        
    except Exception as e:
        print(f"Error in check_budget_balance: {str(e)}")
        return jsonify({'error': f'Error checking budget balance: {str(e)}'}), 500

@app.route('/api/budget/overspending-check', methods=['GET'])
@token_required
def check_overspending(current_user):
    """Check for subcategories where spending exceeds allocation"""
    try:
        # Get active budget period
        active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
        
        if not active_period:
            return jsonify({
                'has_overspending': False,
                'overspent_categories': [],
                'message': 'No active budget period found'
            })
        
        budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
        
        if not budget:
            return jsonify({
                'has_overspending': False,
                'overspent_categories': [],
                'message': 'No budget found for active period'
            })
        
        overspent_categories = []
        
        # Get all subcategories for the user
        subcategories = Subcategory.query.join(Category).filter(Category.user_id == current_user.id).all()
        print(f"DEBUG: Found {len(subcategories)} subcategories for user {current_user.id}")
        
        for subcategory in subcategories:
            # Get allocation for this subcategory
            allocation = BudgetAllocation.query.filter_by(
                budget_id=budget.id,
                subcategory_id=subcategory.id
            ).first()
            
            allocated_amount = allocation.allocated_amount if allocation else 0
            
            # Calculate spent amount for the current period
            spent_transactions = Transaction.query.filter_by(
                user_id=current_user.id,
                subcategory_id=subcategory.id
            ).filter(
                Transaction.transaction_date >= active_period.start_date,
                Transaction.transaction_date <= active_period.end_date
            ).all()
            
            spent_amount = sum(t.amount for t in spent_transactions)
            
            # Check if overspent
            print(f"DEBUG: {subcategory.name}: allocated={allocated_amount}, spent={spent_amount}")
            if spent_amount > allocated_amount:
                overspent_amount = spent_amount - allocated_amount
                print(f"DEBUG: OVERSHOT! {subcategory.name} by {overspent_amount}")
                
                # Calculate percentage (handle division by zero for 0 allocation)
                if allocated_amount > 0:
                    overspent_percentage = (overspent_amount / allocated_amount) * 100
                else:
                    overspent_percentage = 100  # 100% over when allocation is 0
                
                overspent_categories.append({
                    'subcategory_id': subcategory.id,
                    'subcategory_name': subcategory.name,
                    'category_name': subcategory.category.name,
                    'allocated': allocated_amount,
                    'spent': spent_amount,
                    'overspent_amount': overspent_amount,
                    'overspent_percentage': overspent_percentage
                })
        
        # Sort by overspent amount (highest first)
        overspent_categories.sort(key=lambda x: x['overspent_amount'], reverse=True)
        
        print(f"DEBUG: Final result - {len(overspent_categories)} overspent categories")
        for cat in overspent_categories:
            print(f"DEBUG:   - {cat['category_name']} - {cat['subcategory_name']}: {cat['overspent_amount']} over")
        
        return jsonify({
            'has_overspending': len(overspent_categories) > 0,
            'overspent_categories': overspent_categories,
            'total_overspent_categories': len(overspent_categories),
            'message': f'Found {len(overspent_categories)} overspent subcategor{"y" if len(overspent_categories) == 1 else "ies"}'
        })
        
    except Exception as e:
        print(f"Error in check_overspending: {str(e)}")
        return jsonify({'error': f'Error checking overspending: {str(e)}'}), 500


def create_default_categories(user_id):
    default_categories = [
        {
            'name': 'Church and Family',
            'subcategories': ['Tithe', 'Offering', 'Social Responsibility']
        },
        {
            'name': 'Groceries and Food',
            'subcategories': ['Groceries', 'Dining out']
        },
        {
            'name': 'Home Expenses',
            'subcategories': ['Water bill', 'Electricity bill', 'Fibre', 'Rent/Bond repayment']
        },
        {
            'name': 'Monthly Commitments',
            'subcategories': ['Medical Aid', 'Life cover', 'Netflix', 'Education', 'Phone', 'Banking Fees']
        },
        {
            'name': 'Car and Travel',
            'subcategories': ['Car finance', 'Car insurance', 'Car Tracker', 'Fuel', 'Car wash']
        },
        {
            'name': 'Personal Care',
            'subcategories': []
        },
        {
            'name': 'Leisure',
            'subcategories': []
        },
        {
            'name': 'Other',
            'subcategories': []
        }
    ]
    
    for cat_data in default_categories:
        category = Category(name=cat_data['name'], user_id=user_id)
        db.session.add(category)
        db.session.flush()  # Get the category ID
        
        for subcat_name in cat_data['subcategories']:
            subcategory = Subcategory(name=subcat_name, category_id=category.id)
            db.session.add(subcategory)
    
    db.session.commit()

def send_email(to_email, subject, body):
    """Send email using SMTP"""
    try:
        msg = MIMEMultipart()
        msg['From'] = app.config['MAIL_DEFAULT_SENDER']
        msg['To'] = to_email
        msg['Subject'] = subject
        
        msg.attach(MIMEText(body, 'html'))
        
        server = smtplib.SMTP(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        if app.config['MAIL_USE_TLS']:
            server.starttls()
        
        if app.config['MAIL_USERNAME'] and app.config['MAIL_PASSWORD']:
            server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        
        text = msg.as_string()
        server.sendmail(app.config['MAIL_DEFAULT_SENDER'], to_email, text)
        server.quit()
        
        return True
    except Exception as e:
        print(f"Error sending email: {str(e)}")
        return False

def send_verification_email(user, verification_token):
    """Send email verification email"""
    verification_url = f"{request.url_root}verify-email?token={verification_token}"
    
    subject = "Verify Your Email - STEWARD"
    body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="text-align: center; margin-bottom: 30px;">
            <h1 style="color: #8B4513; margin-bottom: 10px;">STEWARD</h1>
            <p style="color: #666; font-size: 16px;">Christian Family Budgeting</p>
        </div>
        
        <div style="background-color: #f8f9fa; padding: 20px; border-radius: 8px; margin-bottom: 20px;">
            <h2 style="color: #333; margin-bottom: 15px;">Welcome to STEWARD!</h2>
            <p style="color: #555; line-height: 1.6; margin-bottom: 15px;">
                Hi {user.preferred_name or user.first_name},
            </p>
            <p style="color: #555; line-height: 1.6; margin-bottom: 20px;">
                Thank you for signing up for STEWARD! To complete your account setup and start managing your finances, 
                please verify your email address by clicking the button below.
            </p>
            
            <div style="text-align: center; margin: 25px 0;">
                <a href="{verification_url}" 
                   style="background-color: #8B4513; color: white; padding: 12px 30px; text-decoration: none; 
                          border-radius: 5px; font-weight: bold; display: inline-block;">
                    Verify Email Address
                </a>
            </div>
            
            <p style="color: #666; font-size: 14px; margin-top: 20px;">
                If the button doesn't work, you can copy and paste this link into your browser:<br>
                <a href="{verification_url}" style="color: #8B4513; word-break: break-all;">{verification_url}</a>
            </p>
        </div>
        
        <div style="text-align: center; color: #666; font-size: 14px; margin-top: 30px;">
            <p>This verification link will expire in 24 hours.</p>
            <p>If you didn't create an account with STEWARD, please ignore this email.</p>
        </div>
        
        <div style="text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee;">
            <p style="color: #8B4513; font-style: italic; margin: 0;">
                "The plans of the diligent lead to profit as surely as haste leads to poverty." - Proverbs 21:5
            </p>
        </div>
    </body>
    </html>
    """
    
    return send_email(user.email, subject, body)

# Onboarding Routes
@app.route('/onboarding')
def onboarding():
    """Onboarding page for new users"""
    return render_template('onboarding.html')

@app.route('/api/validate-email', methods=['POST'])
def validate_email():
    """Validate if email already exists"""
    try:
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({'exists': False, 'message': 'Email is required'}), 400
        
        # Check if email already exists
        existing_user = User.query.filter_by(email=email).first()
        
        return jsonify({
            'exists': existing_user is not None,
            'message': 'Email already exists' if existing_user else 'Email is available'
        })
        
    except Exception as e:
        print(f"Email validation error: {str(e)}")
        return jsonify({'exists': False, 'message': 'Validation error'}), 500

@app.route('/api/validate-username', methods=['POST'])
def validate_username():
    """Validate if username already exists"""
    try:
        data = request.get_json()
        username = data.get('username')
        
        if not username:
            return jsonify({'exists': False, 'message': 'Username is required'}), 400
        
        # Validate username format
        import re
        if not re.match(r'^[a-zA-Z0-9_]{3,20}$', username):
            return jsonify({'exists': True, 'message': 'Username must be 3-20 characters, letters, numbers, and underscores only'}), 400
        
        # Check if username already exists
        existing_user = User.query.filter_by(username=username).first()
        
        return jsonify({
            'exists': existing_user is not None,
            'message': 'Username already exists' if existing_user else 'Username is available'
        })
        
    except Exception as e:
        print(f"Username validation error: {str(e)}")
        return jsonify({'exists': False, 'message': 'Validation error'}), 500

@app.route('/api/onboarding/complete', methods=['POST'])
def complete_onboarding():
    """Complete the onboarding process and create user account"""
    try:
        data = request.get_json()
        
        # Debug: Log received data
        print("=== ONBOARDING DATA RECEIVED ===")
        print("Categories:", data.get('categories', []))
        print("Subcategories:", data.get('subcategories', []))
        print("Full data keys:", list(data.keys()))
        print("================================")
        
        # Extract form data
        personal_info = {
            'firstName': data.get('firstName'),
            'lastName': data.get('lastName'),
            'email': data.get('email'),
            'username': data.get('username'),
            'country': data.get('country'),
            'preferredName': data.get('preferredName')
        }
        
        password_info = {
            'password': data.get('password'),
            'confirmPassword': data.get('confirmPassword')
        }
        
        referral_info = {
            'referralSource': data.get('referralSource'),
            'referralDetailsText': data.get('referralDetailsText')
        }
        
        details_info = {
            'currency': data.get('currency'),
            'categories': data.get('categories', []),
            'subcategories': data.get('subcategories', [])
        }
        
        # Validate required fields
        required_fields = ['firstName', 'lastName', 'email', 'password', 'currency']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'{field} is required'}), 400
        
        # Validate password match
        if password_info['password'] != password_info['confirmPassword']:
            return jsonify({'success': False, 'message': 'Passwords do not match'}), 400
        
        # Check if user already exists
        existing_user = User.query.filter_by(email=personal_info['email']).first()
        if existing_user:
            return jsonify({'success': False, 'message': 'An account with this email already exists'}), 400
        
        # Create username from provided username, preferred name, or email
        provided_username = personal_info.get('username', '').strip()
        if provided_username:
            # Validate username format if provided
            import re
            if not re.match(r'^[a-zA-Z0-9_]{3,20}$', provided_username):
                return jsonify({'success': False, 'message': 'Username must be 3-20 characters, letters, numbers, and underscores only'}), 400
            username = provided_username
        else:
            username = personal_info.get('preferredName') or personal_info['email'].split('@')[0]
        
        # Ensure username is unique
        counter = 1
        original_username = username
        while User.query.filter_by(username=username).first():
            username = f"{original_username}{counter}"
            counter += 1
        
        # Create new user
        user = User(
            username=username,
            email=personal_info['email'],
            first_name=personal_info['firstName'],
            last_name=personal_info['lastName'],
            country=personal_info['country'],
            preferred_name=personal_info['preferredName'],
            currency=details_info['currency'],
            referral_source=referral_info['referralSource'],
            referral_details=referral_info['referralDetailsText']
        )
        user.set_password(password_info['password'])
        
        db.session.add(user)
        db.session.commit()
        
        # Create categories and subcategories based on user selection
        category_mapping = {
            'faithful-stewardship': {
                'name': 'Giving',
                'subcategories': {
                    'tithe': 'Tithe',
                    'offering': 'Offering',
                    'social-responsibility': 'Social Responsibility'
                }
            },
            'groceries': {
                'name': 'Groceries',
                'subcategories': {
                    'food-home-essentials': 'Food & Home Essentials',
                    'dining-out': 'Dining out'
                }
            },
            'housing': {
                'name': 'Housing',
                'subcategories': {
                    'mortgage-rent': 'Mortgage/Rent',
                    'hoa-fees-levies': 'HOA Fees/Levies',
                    'electricity-bill': 'Electricity Bill',
                    'water-bill': 'Water Bill',
                    'home-maintenance': 'Home maintenance',
                    'home-insurance': 'Home Insurance',
                    'internet': 'Internet'
                }
            },
            'transportation': {
                'name': 'Transportation',
                'subcategories': {
                    'loan-repayment': 'Loan repayment',
                    'insurance': 'Insurance',
                    'fuel': 'Fuel',
                    'car-tracker': 'Car Tracker',
                    'car-wash': 'Car wash'
                }
            },
            'monthly-commitments': {
                'name': 'Monthly Commitments',
                'subcategories': {
                    'life-cover': 'Life cover',
                    'funeral-plan': 'Funeral Plan',
                    'credit-card-repayment': 'Credit card repayment',
                    'monthly-banking-fees': 'Monthly Banking Fees'
                }
            },
            'leisure-entertainment': {
                'name': 'Leisure/Entertainment',
                'subcategories': {
                    'spotify': 'Spotify',
                    'weekend-adventures': 'Weekend adventures'
                }
            },
            'personal-care': {
                'name': 'Personal Care',
                'subcategories': {
                    'gym-membership': 'Gym membership',
                    'haircuts': 'Haircuts',
                    'clothing': 'Clothing'
                }
            },
            'savings-goals': {
                'name': 'Savings Goals',
                'subcategories': {
                    'emergency-fund': 'Emergency fund',
                    'general-savings': 'General Savings',
                    'short-term-goal': 'Short term goal'
                }
            },
            'once-off-expenses': {
                'name': 'Once-off expenses (populated as it happens)',
                'subcategories': {
                    'asset-purchase': 'Asset purchase',
                    'emergency': 'Emergency'
                }
            }
        }
        
        selected_categories = details_info.get('categories', [])
        selected_subcategories = details_info.get('subcategories', [])
        custom_subcategory_names = data.get('custom_subcategory_names', {})
        
        # Process custom categories and their subcategories together
        print(f"All selected subcategories: {selected_subcategories}")
        custom_subcategories = [sub for sub in selected_subcategories if sub.startswith('custom-subcategory-')]
        print(f"Found custom subcategories: {custom_subcategories}")
        
        # Group custom subcategories by their parent category
        custom_category_subcategories = {}
        for subcategory_key in custom_subcategories:
            # Extract the parent category ID from the subcategory key
            # Format: custom-subcategory-{parent_category_id}-{subcategory_counter}
            # Examples: 
            # - custom-subcategory-custom-category-1-2 (for custom categories)
            # - custom-subcategory-faithful-stewardship-1 (for predefined categories)
            parts = subcategory_key.split('-')
            if len(parts) >= 4:  # custom-subcategory-{parent}-{counter}
                if parts[2] == 'custom' and len(parts) >= 5:  # custom-subcategory-custom-category-{number}-{counter}
                    parent_category_id = f"{parts[2]}-{parts[3]}-{parts[4]}"  # Extract custom-category-{number}
                else:  # custom-subcategory-{predefined_category}-{counter}
                    parent_category_id = parts[2]  # Extract the predefined category ID
                
                if parent_category_id not in custom_category_subcategories:
                    custom_category_subcategories[parent_category_id] = []
                custom_category_subcategories[parent_category_id].append(subcategory_key)
        
        print(f"Custom category subcategories mapping: {custom_category_subcategories}")
        print(f"Selected categories: {selected_categories}")
        
        for category_key in selected_categories:
            print(f"Processing category: {category_key}")
            if category_key in category_mapping:
                # Handle predefined categories
                category_data = category_mapping[category_key]
                print(f"Category data: {category_data}")
                category = Category(
                    name=category_data['name'],
                    user_id=user.id,
                    is_template=True
                )
                db.session.add(category)
                db.session.flush()  # Get the category ID
                print(f"Created category with ID: {category.id}")
                
                # Add selected subcategories for this category
                subcategories_added = 0
                for subcategory_key in selected_subcategories:
                    print(f"Checking subcategory: {subcategory_key}")
                    if subcategory_key in category_data['subcategories']:
                        subcategory = Subcategory(
                            name=category_data['subcategories'][subcategory_key],
                            category_id=category.id
                        )
                        db.session.add(subcategory)
                        subcategories_added += 1
                        print(f"Added subcategory: {subcategory_key} -> {category_data['subcategories'][subcategory_key]}")
                    else:
                        print(f"Subcategory {subcategory_key} not found in category {category_key}")
                
                # Also add custom subcategories for this predefined category
                if category_key in custom_category_subcategories:
                    print(f"Found custom subcategories for predefined category {category_key}: {custom_category_subcategories[category_key]}")
                    for subcategory_key in custom_category_subcategories[category_key]:
                        # Get the actual subcategory name from the frontend
                        subcategory_name = custom_subcategory_names.get(subcategory_key, f"Custom Subcategory {subcategory_key.split('-')[-1]}")
                        subcategory = Subcategory(
                            name=subcategory_name,
                            category_id=category.id
                        )
                        db.session.add(subcategory)
                        subcategories_added += 1
                        print(f"Added custom subcategory to predefined category: {subcategory_key} -> {subcategory_name}")
                
                print(f"Total subcategories added for {category_key}: {subcategories_added}")
            elif category_key.startswith('custom-category-'):
                # Handle custom categories
                print(f"Processing custom category: {category_key}")
                
                # Get the custom category name from the form data
                custom_category_name = data.get('custom_category_names', {}).get(category_key, f'Custom Category {category_key.split("-")[-1]}')
                print(f"Custom category name: {custom_category_name}")
                
                category = Category(
                    name=custom_category_name,
                    user_id=user.id,
                    is_template=False
                )
                db.session.add(category)
                db.session.flush()  # Get the category ID
                print(f"Created custom category with ID: {category.id}")
                
                # Add subcategories for this custom category
                subcategories_added = 0
                print(f"Looking for subcategories for category {category_key} in mapping: {custom_category_subcategories}")
                if category_key in custom_category_subcategories:
                    print(f"Found subcategories for {category_key}: {custom_category_subcategories[category_key]}")
                    for subcategory_key in custom_category_subcategories[category_key]:
                        # Get the actual subcategory name from the frontend
                        subcategory_name = custom_subcategory_names.get(subcategory_key, f"Custom Subcategory {subcategory_key.split('-')[-1]}")
                        subcategory = Subcategory(
                            name=subcategory_name,
                            category_id=category.id
                        )
                        db.session.add(subcategory)
                        subcategories_added += 1
                        print(f"Added custom subcategory: {subcategory_key} -> {subcategory_name}")
                else:
                    print(f"No subcategories found for category {category_key}")
                
                print(f"Total subcategories added for custom category {category_key}: {subcategories_added}")
            else:
                print(f"Category {category_key} not found in mapping and not a custom category")
        
        db.session.commit()
        
        # Create email verification token
        verification_token = secrets.token_urlsafe(32)
        verification = EmailVerification(
            user_id=user.id,
            token=verification_token,
            expires_at=datetime.utcnow() + timedelta(hours=24)
        )
        db.session.add(verification)
        db.session.commit()
        
        # Send verification email
        email_sent = send_verification_email(user, verification_token)
        
        if email_sent:
            return jsonify({
                'success': True,
                'message': 'Account created successfully! Please check your email to verify your account.',
                'email_verification_required': True,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'currency': user.currency
                }
            })
        else:
            # If email fails, still create account but mark as unverified
            return jsonify({
                'success': True,
                'message': 'Account created successfully! Please check your email to verify your account.',
                'email_verification_required': True,
                'email_sent': False,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'currency': user.currency
                }
            })
        
    except Exception as e:
        db.session.rollback()
        print(f"Onboarding error: {str(e)}")
        return jsonify({'success': False, 'message': 'An error occurred during account creation'}), 500

@app.route('/verify-email')
def verify_email():
    """Email verification page"""
    token = request.args.get('token')
    if not token:
        return render_template('email_verification.html', 
                             success=False, 
                             message='Invalid verification link')
    
    # Find verification token
    verification = EmailVerification.query.filter_by(token=token, verified=False).first()
    
    if not verification:
        return render_template('email_verification.html', 
                             success=False, 
                             message='Invalid or expired verification link')
    
    # Check if token is expired
    if verification.expires_at < datetime.utcnow():
        return render_template('email_verification.html', 
                             success=False, 
                             message='Verification link has expired. Please request a new one.')
    
    # Verify the user's email
    user = User.query.get(verification.user_id)
    if user:
        user.email_verified = True
        verification.verified = True
        db.session.commit()
        
        return render_template('email_verification.html', 
                             success=True, 
                             message='Email verified successfully! You can now log in to your account.')
    else:
        return render_template('email_verification.html', 
                             success=False, 
                             message='User not found')

@app.route('/api/resend-verification', methods=['POST'])
def resend_verification():
    """Resend verification email"""
    try:
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({'success': False, 'message': 'Email is required'}), 400
        
        # Find user by email
        user = User.query.filter_by(email=email).first()
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404
        
        if user.email_verified:
            return jsonify({'success': False, 'message': 'Email already verified'}), 400
        
        # Create new verification token
        verification_token = secrets.token_urlsafe(32)
        verification = EmailVerification(
            user_id=user.id,
            token=verification_token,
            expires_at=datetime.utcnow() + timedelta(hours=24)
        )
        db.session.add(verification)
        db.session.commit()
        
        # Send verification email
        email_sent = send_verification_email(user, verification_token)
        
        if email_sent:
            return jsonify({'success': True, 'message': 'Verification email sent successfully'})
        else:
            return jsonify({'success': False, 'message': 'Failed to send verification email'}), 500
            
    except Exception as e:
        print(f"Error resending verification: {str(e)}")
        return jsonify({'success': False, 'message': 'An error occurred'}), 500

@app.route('/api/user/theme', methods=['POST'])
@token_required
def update_user_theme(current_user):
    """Update user theme preference"""
    try:
        data = request.get_json()
        theme = data.get('theme', 'dark')
        
        if theme not in ['light', 'dark']:
            return jsonify({'success': False, 'message': 'Invalid theme'}), 400
        
        # Update user theme in database
        current_user.theme = theme
        db.session.commit()
        return jsonify({'success': True, 'message': 'Theme updated successfully'})
            
    except Exception as e:
        print(f'Error updating theme: {e}')
        return jsonify({'success': False, 'message': 'Failed to update theme'}), 500

# Helper function to populate budget from recurring sources
def populate_budget_from_recurring(user, budget):
    """Populate a new budget with recurring income sources and allocations"""
    try:
        # Check if budget already has data to avoid duplicates
        existing_income_sources = IncomeSource.query.filter_by(budget_id=budget.id).count()
        existing_allocations = BudgetAllocation.query.filter_by(budget_id=budget.id).count()
        
        if existing_income_sources > 0 or existing_allocations > 0:
            print(f"Budget {budget.id} already has data, skipping auto-population to avoid duplicates")
            return
        
        # Add recurring income sources
        recurring_income_sources = RecurringIncomeSource.query.filter_by(
            user_id=user.id, 
            is_active=True
        ).all()
        
        for recurring_source in recurring_income_sources:
            income_source = IncomeSource(
                name=recurring_source.name,
                amount=recurring_source.amount,
                budget_id=budget.id
            )
            db.session.add(income_source)
        
        # Add recurring budget allocations
        recurring_allocations = RecurringBudgetAllocation.query.filter_by(
            user_id=user.id, 
            is_active=True
        ).all()
        
        for recurring_allocation in recurring_allocations:
            budget_allocation = BudgetAllocation(
                allocated_amount=recurring_allocation.allocated_amount,
                subcategory_id=recurring_allocation.subcategory_id,
                budget_id=budget.id
            )
            db.session.add(budget_allocation)
        
        # Update total income
        budget.total_income = sum(source.amount for source in budget.income_sources)
        
        db.session.commit()
        
    except Exception as e:
        print(f"Error populating budget from recurring sources: {str(e)}")
        db.session.rollback()

# Recurring Income Sources API Endpoints
@app.route('/api/recurring-income-sources', methods=['GET'])
@token_required
def get_recurring_income_sources(current_user):
    """Get all recurring income sources for the user"""
    try:
        recurring_sources = RecurringIncomeSource.query.filter_by(
            user_id=current_user.id, 
            is_active=True
        ).order_by(RecurringIncomeSource.name).all()
        
        return jsonify([{
            'id': source.id,
            'name': source.name,
            'amount': source.amount,
            'created_at': source.created_at.isoformat() if source.created_at else None
        } for source in recurring_sources])
    except Exception as e:
        return jsonify({'error': f'Error loading recurring income sources: {str(e)}'}), 500

@app.route('/api/recurring-income-sources', methods=['POST'])
@token_required
def create_recurring_income_source(current_user):
    """Create a new recurring income source"""
    try:
        data = request.get_json()
        
        recurring_source = RecurringIncomeSource(
            name=data['name'],
            amount=float(data['amount']),
            user_id=current_user.id
        )
        
        db.session.add(recurring_source)
        db.session.commit()
        
        return jsonify({
            'message': 'Recurring income source created successfully',
            'id': recurring_source.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error creating recurring income source: {str(e)}'}), 500

@app.route('/api/recurring-income-sources/<int:source_id>', methods=['PUT'])
@token_required
def update_recurring_income_source(current_user, source_id):
    """Update a recurring income source"""
    try:
        data = request.get_json()
        
        recurring_source = RecurringIncomeSource.query.filter_by(
            id=source_id, 
            user_id=current_user.id
        ).first()
        
        if not recurring_source:
            return jsonify({'message': 'Recurring income source not found'}), 404
        
        recurring_source.name = data['name']
        recurring_source.amount = float(data['amount'])
        recurring_source.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'message': 'Recurring income source updated successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error updating recurring income source: {str(e)}'}), 500

@app.route('/api/recurring-income-sources/<int:source_id>', methods=['DELETE'])
@token_required
def delete_recurring_income_source(current_user, source_id):
    """Delete a recurring income source"""
    try:
        recurring_source = RecurringIncomeSource.query.filter_by(
            id=source_id, 
            user_id=current_user.id
        ).first()
        
        if not recurring_source:
            return jsonify({'message': 'Recurring income source not found'}), 404
        
        recurring_source.is_active = False
        recurring_source.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'message': 'Recurring income source deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error deleting recurring income source: {str(e)}'}), 500

# Recurring Budget Allocations API Endpoints
@app.route('/api/recurring-allocations', methods=['GET'])
@token_required
def get_recurring_allocations(current_user):
    """Get all recurring budget allocations for the user"""
    try:
        recurring_allocations = RecurringBudgetAllocation.query.filter_by(
            user_id=current_user.id, 
            is_active=True
        ).join(Subcategory).join(Category).order_by(Category.name, Subcategory.name).all()
        
        return jsonify([{
            'id': allocation.id,
            'allocated_amount': allocation.allocated_amount,
            'subcategory_id': allocation.subcategory_id,
            'subcategory_name': allocation.subcategory.name,
            'category_name': allocation.subcategory.category.name,
            'created_at': allocation.created_at.isoformat() if allocation.created_at else None
        } for allocation in recurring_allocations])
    except Exception as e:
        return jsonify({'error': f'Error loading recurring allocations: {str(e)}'}), 500

@app.route('/api/recurring-allocations', methods=['POST'])
@token_required
def create_recurring_allocation(current_user):
    """Create a new recurring budget allocation"""
    try:
        data = request.get_json()
        
        # Verify subcategory exists and belongs to user
        subcategory = Subcategory.query.join(Category).filter(
            Subcategory.id == data['subcategory_id'],
            Category.user_id == current_user.id
        ).first()
        
        if not subcategory:
            return jsonify({'message': 'Subcategory not found'}), 404
        
        recurring_allocation = RecurringBudgetAllocation(
            allocated_amount=float(data['allocated_amount']),
            subcategory_id=data['subcategory_id'],
            user_id=current_user.id
        )
        
        db.session.add(recurring_allocation)
        db.session.commit()
        
        return jsonify({
            'message': 'Recurring allocation created successfully',
            'id': recurring_allocation.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error creating recurring allocation: {str(e)}'}), 500

@app.route('/api/recurring-allocations/<int:allocation_id>', methods=['PUT'])
@token_required
def update_recurring_allocation(current_user, allocation_id):
    """Update a recurring budget allocation"""
    try:
        data = request.get_json()
        
        recurring_allocation = RecurringBudgetAllocation.query.filter_by(
            id=allocation_id, 
            user_id=current_user.id
        ).first()
        
        if not recurring_allocation:
            return jsonify({'message': 'Recurring allocation not found'}), 404
        
        recurring_allocation.allocated_amount = float(data['allocated_amount'])
        recurring_allocation.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'message': 'Recurring allocation updated successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error updating recurring allocation: {str(e)}'}), 500

@app.route('/api/recurring-allocations/<int:allocation_id>', methods=['DELETE'])
@token_required
def delete_recurring_allocation(current_user, allocation_id):
    """Delete a recurring budget allocation"""
    try:
        recurring_allocation = RecurringBudgetAllocation.query.filter_by(
            id=allocation_id, 
            user_id=current_user.id
        ).first()
        
        if not recurring_allocation:
            return jsonify({'message': 'Recurring allocation not found'}), 404
        
        recurring_allocation.is_active = False
        recurring_allocation.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'message': 'Recurring allocation deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error deleting recurring allocation: {str(e)}'}), 500

# Manual trigger for auto-populating current budget from recurring sources
@app.route('/api/budget/populate-recurring', methods=['POST'])
@token_required
def populate_current_budget_from_recurring(current_user):
    """Manually populate the current active budget with recurring sources"""
    try:
        # Get active budget period
        active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
        if not active_period:
            return jsonify({'message': 'No active budget period found'}), 404
        
        # Get budget for this period
        budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
        if not budget:
            return jsonify({'message': 'No budget found for active period'}), 404
        
        # Populate from recurring sources
        populate_budget_from_recurring(current_user, budget)
        
        return jsonify({'message': 'Budget populated with recurring sources successfully'})
        
    except Exception as e:
        return jsonify({'message': f'Error populating budget: {str(e)}'}), 500

# Cleanup endpoint to remove duplicate allocations
@app.route('/api/budget/cleanup-duplicates', methods=['POST'])
@token_required
def cleanup_duplicate_allocations(current_user):
    """Remove duplicate allocations from the current budget"""
    try:
        # Get active budget period
        active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
        if not active_period:
            return jsonify({'message': 'No active budget period found'}), 404
        
        budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
        if not budget:
            return jsonify({'message': 'No budget found for active period'}), 404
        
        # Get all allocations for this budget
        allocations = BudgetAllocation.query.filter_by(budget_id=budget.id).all()
        
        # Group by subcategory_id and keep only the first one
        seen_subcategories = set()
        duplicates_removed = 0
        
        for allocation in allocations:
            if allocation.subcategory_id in seen_subcategories:
                # This is a duplicate, remove it
                db.session.delete(allocation)
                duplicates_removed += 1
            else:
                seen_subcategories.add(allocation.subcategory_id)
        
        db.session.commit()
        
        return jsonify({
            'message': f'Cleanup completed. Removed {duplicates_removed} duplicate allocations.',
            'duplicates_removed': duplicates_removed
        })
        
    except Exception as e:
        return jsonify({'message': f'Error cleaning up duplicates: {str(e)}'}), 500

# Debug endpoint to help diagnose budget calculation issues
@app.route('/api/debug/budget-calculation', methods=['GET'])
@token_required
def debug_budget_calculation(current_user):
    """Debug endpoint to help diagnose budget calculation issues"""
    try:
        # Get active budget period
        active_period = BudgetPeriod.query.filter_by(user_id=current_user.id, is_active=True).first()
        if not active_period:
            return jsonify({'error': 'No active budget period found'})
        
        budget = Budget.query.filter_by(period_id=active_period.id, user_id=current_user.id).first()
        if not budget:
            return jsonify({'error': 'No budget found for active period'})
        
        # Get all budgets for this user (for comparison)
        all_budgets = Budget.query.filter_by(user_id=current_user.id).all()
        
        # Get all allocations for this user (for comparison)
        all_allocations = BudgetAllocation.query.join(Budget).filter(Budget.user_id == current_user.id).all()
        
        # Calculate totals for current budget
        current_income = sum(source.amount for source in budget.income_sources)
        current_balance_forward = budget.balance_brought_forward or 0
        current_allocated = sum(allocation.allocated_amount for allocation in budget.allocations)
        
        # Calculate totals across all budgets
        total_income_all_budgets = sum(
            sum(source.amount for source in b.income_sources) + (b.balance_brought_forward or 0)
            for b in all_budgets
        )
        total_allocated_all_budgets = sum(
            sum(allocation.allocated_amount for allocation in b.allocations)
            for b in all_budgets
        )
        
        return jsonify({
            'current_budget': {
                'budget_id': budget.id,
                'period_name': active_period.name,
                'income_sources_count': len(budget.income_sources),
                'allocations_count': len(budget.allocations),
                'total_income': current_income,
                'balance_forward': current_balance_forward,
                'total_available': current_income + current_balance_forward,
                'total_allocated': current_allocated,
                'balance': (current_income + current_balance_forward) - current_allocated
            },
            'all_budgets_summary': {
                'total_budgets': len(all_budgets),
                'total_income_all': total_income_all_budgets,
                'total_allocated_all': total_allocated_all_budgets,
                'balance_all': total_income_all_budgets - total_allocated_all_budgets
            },
            'allocations_detail': [
                {
                    'budget_id': allocation.budget_id,
                    'subcategory_id': allocation.subcategory_id,
                    'amount': allocation.allocated_amount,
                    'period_name': allocation.budget.period.name if allocation.budget.period else 'Unknown'
                }
                for allocation in all_allocations
            ]
        })
        
    except Exception as e:
        return jsonify({'error': f'Debug error: {str(e)}'}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000)
